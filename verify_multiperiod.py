import openpyxl

wb = openpyxl.load_workbook('output/financial_statements/MSFT_789019_financial_statements.xlsx')
ws = wb['Balance Sheet']

print("Balance Sheet - First 15 rows showing ALL PERIODS:")
print("=" * 120)

# Show first 15 rows, columns A through G (to show both periods in both sections)
for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=7, values_only=True):
    formatted = []
    for i, cell in enumerate(row):
        if cell is None:
            formatted.append(' ' * 18)
        elif isinstance(cell, (int, float)):
            formatted.append(f"{cell:>18,.0f}")
        else:
            formatted.append(f"{str(cell)[:18]:<18}")
    print(" | ".join(formatted))

print("\n" + "=" * 120)
print("Checking Income Statement periods...")
ws = wb['Income Statement']
# Show header rows
for row in ws.iter_rows(min_row=6, max_row=8, min_col=1, max_col=7, values_only=True):
    formatted = []
    for cell in row:
        if cell is None:
            formatted.append(' ' * 25)
        elif isinstance(cell, (int, float)):
            formatted.append(f"{cell:>25,.0f}")
        else:
            formatted.append(f"{str(cell)[:25]:<25}")
    print(" | ".join(formatted))
