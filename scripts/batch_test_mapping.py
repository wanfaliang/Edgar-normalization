"""
Batch Test Financial Statement Mapping
======================================
Test financial statement mapping across companies for a specific fiscal period.
Supports testing predefined company groups or specific tickers/CIKs.

Usage:
    # Test all 50 general companies for 2024 Q2
    python scripts/batch_test_mapping.py --year 2024 --quarter 2 --group general

    # Test all 25 financial companies for 2024 Q2
    python scripts/batch_test_mapping.py --year 2024 --quarter 2 --group financial

    # Test specific companies by ticker
    python scripts/batch_test_mapping.py --year 2024 --quarter 2 --ticker MSFT AAPL AMZN

    # Test specific companies by CIK
    python scripts/batch_test_mapping.py --year 2024 --quarter 2 --cik 789019 320193

    # Filter by form type (10-Q or 10-K)
    python scripts/batch_test_mapping.py --year 2024 --quarter 2 --group general --form 10-K

Required Arguments:
    --year YYYY         Fiscal year (e.g., 2024)
    --quarter {1,2,3,4} Fiscal quarter

Optional Arguments:
    --group {general,financial}  Predefined company group to test
    --form {10-Q,10-K}           Filter by form type
    --ticker [TICKERS]           Test specific tickers (space-separated)
    --cik [CIKS]                 Test specific CIKs (space-separated)

Output:
    output/results/mapping_test_TIMESTAMP.xlsx        Test results matrix
    output/results/mapping_test_TIMESTAMP/            Folder with individual company statements
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from statement_reconstructor import StatementReconstructor
from map_financial_statements import (
    find_bs_control_items,
    find_is_control_items,
    find_cf_control_items,
    map_financial_statements,
    create_excel_workbook
)
import psycopg2
from psycopg2.extras import RealDictCursor
from config import config
import argparse

# =============================================================================
# PREDEFINED COMPANY GROUPS
# =============================================================================

# 50 General Companies (diverse industries)
GENERAL_COMPANIES = [
    789019, 320193, 1652044, 1018724, 1326801, 1045810, 50863, 1341439, 1108524, 796343,
    104169, 27419, 909832, 354950, 60667, 109198, 29534, 764478,
    66740, 18230, 12927, 936468, 315189, 40545, 773840,
    19617, 70858, 72971, 886982, 895421, 4962, 1403161, 1141391,
    200406, 78003, 731766, 1800, 310158, 64803,
    34088, 93410, 1163165, 87347,
    21344, 77476, 80424, 320187, 63908, 829224, 21665
]

# 25 Financial Companies (SIC 6000-6799) - banks, insurance, REITs, etc.
FINANCIAL_COMPANIES = [
    1000209,   # MBNKO  - MEDALLION FINANCIAL CORP (SIC 6199)
    1000275,   # RYPBF  - ROYAL BANK OF CANADA (SIC 6029)
    1001171,   # BYFC   - BROADWAY FINANCIAL CORP DE (SIC 6035)
    1001290,   # BAP    - CREDICORP LTD (SIC 6029)
    1004434,   # MGRE   - AFFILIATED MANAGERS GROUP, INC. (SIC 6282)
    1004702,   # OCFC   - OCEANFIRST FINANCIAL CORP (SIC 6021)
    1004724,   # RHEPZ  - REGIONAL HEALTH PROPERTIES, INC (SIC 6519)
    1005101,   # MGLD   - MARYGOLD COMPANIES, INC. (SIC 6199)
    1005817,   # TMP    - TOMPKINS FINANCIAL CORP (SIC 6022)
    1006830,   # CBKM   - CONSUMERS BANCORP INC /OH/ (SIC 6021)
    1010470,   # PROV   - PROVIDENT FINANCIAL HOLDINGS INC (SIC 6035)
    101199,    # UFCS   - UNITED FIRE GROUP INC (SIC 6331)
    1013272,   # NWFL   - NORWOOD FINANCIAL CORP (SIC 6022)
    101382,    # UMBFO  - UMB FINANCIAL CORP (SIC 6021)
    1015328,   # WTFCN  - WINTRUST FINANCIAL CORP (SIC 6022)
    1018979,   # AMSF   - AMERISAFE INC (SIC 6331)
    1020569,   # IRM    - IRON MOUNTAIN INC (SIC 6798)
    1021917,   # AWCA   - AWAYSIS CAPITAL, INC. (SIC 6512)
    102212,    # UVSP   - UNIVEST FINANCIAL CORP (SIC 6022)
    1022837,   # SMFNF  - SUMITOMO MITSUI FINANCIAL GROUP, INC. (SIC 6029)
    1025378,   # WPC    - W. P. CAREY INC. (SIC 6798)
    1025835,   # EFSCP  - ENTERPRISE FINANCIAL SERVICES CORP (SIC 6022)
    1025996,   # KRC    - KILROY REALTY CORP (SIC 6798)
    1026214,   # FREJP  - FEDERAL HOME LOAN MORTGAGE CORP (SIC 6111)
    1029199,   # EEFT   - EURONET WORLDWIDE, INC. (SIC 6099)
]

# =============================================================================
# CONTROL ITEM DEFINITIONS
# =============================================================================

CONTROL_ITEMS = {
    'BS': [
        ('total_current_assets', True),
        ('total_non_current_assets', False),
        ('total_assets', True),
        ('total_current_liabilities', True),
        ('total_liabilities', False),
        ('total_stockholders_equity', True),
        ('total_equity', False),
        ('total_liabilities_and_total_equity', True)
    ],
    'IS': [
        ('revenue', False),
        ('operating_income', False),
        ('income_tax_expense', True),
        ('net_income', True),
        ('eps', True),
        ('eps_diluted', True),
        ('weighted_average_shares_outstanding', False),
        ('weighted_average_shares_outstanding_diluted', False)
    ],
    'CF': [
        ('net_income', True),
        ('net_cash_provided_by_operating_activities', True),
        ('net_cash_provided_by_investing_activities', True),
        ('net_cash_provided_by_financing_activities', True),
        ('cash_at_beginning_of_period', True),
        ('cash_at_end_of_period', True)
    ]
}


def test_company(company_info, reconstructor):
    """Test control items for a single company"""
    result = {
        'company': company_info,
        'BS': {'control_items': {}, 'error': None},
        'IS': {'control_items': {}, 'error': None},
        'CF': {'control_items': {}, 'error': None}
    }

    cik = company_info['cik']
    adsh = company_info['adsh']

    # Test Balance Sheet
    try:
        bs_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type='BS')
        if bs_result and bs_result.get('line_items'):
            control_items = find_bs_control_items(bs_result['line_items'])
            result['BS']['control_items'] = control_items
            result['BS']['line_items'] = bs_result['line_items']
    except Exception as e:
        result['BS']['error'] = str(e)

    # Test Income Statement
    try:
        is_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type='IS')
        if is_result and is_result.get('line_items'):
            control_items = find_is_control_items(is_result['line_items'])
            result['IS']['control_items'] = control_items
            result['IS']['line_items'] = is_result['line_items']
    except Exception as e:
        result['IS']['error'] = str(e)

    # Test Cash Flow
    try:
        cf_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type='CF')
        if cf_result and cf_result.get('line_items'):
            control_items = find_cf_control_items(cf_result['line_items'])
            result['CF']['control_items'] = control_items
            result['CF']['line_items'] = cf_result['line_items']
    except Exception as e:
        result['CF']['error'] = str(e)

    return result


def create_excel_report(results, output_file):
    """Create Excel report with results matrix"""
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    # =========================================================================
    # SHEET 1: SUMMARY DASHBOARD
    # =========================================================================
    ws_summary = wb.create_sheet("Summary")

    ws_summary['A1'] = "Financial Statement Mapping Test Report"
    ws_summary['A1'].font = Font(bold=True, size=14)

    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A3'] = f"Total Companies Tested: {len(results)}"

    # Calculate statistics
    stats = {
        'BS': {'total': 0, 'required': 0, 'total_req': 0},
        'IS': {'total': 0, 'required': 0, 'total_req': 0},
        'CF': {'total': 0, 'required': 0, 'total_req': 0}
    }

    for result in results:
        for stmt_type in ['BS', 'IS', 'CF']:
            control_items = result[stmt_type]['control_items']
            for item_name, is_required in CONTROL_ITEMS[stmt_type]:
                if item_name in control_items:
                    stats[stmt_type]['total'] += 1
                    if is_required:
                        stats[stmt_type]['required'] += 1
                if is_required:
                    stats[stmt_type]['total_req'] += 1

    row = 5
    ws_summary[f'A{row}'] = "Overall Success Rates"
    ws_summary[f'A{row}'].font = bold_font
    row += 1

    for stmt_type, name in [('BS', 'Balance Sheet'), ('IS', 'Income Statement'), ('CF', 'Cash Flow')]:
        total_items = len(CONTROL_ITEMS[stmt_type]) * len(results)
        found_items = stats[stmt_type]['total']
        pct = (found_items / total_items * 100) if total_items > 0 else 0

        required_found = stats[stmt_type]['required']
        required_total = stats[stmt_type]['total_req']
        req_pct = (required_found / required_total * 100) if required_total > 0 else 0

        ws_summary[f'A{row}'] = f"{name}:"
        ws_summary[f'B{row}'] = f"{found_items}/{total_items} ({pct:.1f}%)"
        ws_summary[f'C{row}'] = f"Required: {required_found}/{required_total} ({req_pct:.1f}%)"
        row += 1

    # =========================================================================
    # SHEETS 2-4: STATEMENT MATRICES (BS, IS, CF)
    # =========================================================================
    for stmt_type, stmt_name in [('BS', 'Balance Sheet'), ('IS', 'Income Statement'), ('CF', 'Cash Flow')]:
        ws = wb.create_sheet(stmt_name)

        # Headers
        ws['A1'] = "Company"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'] = "Ticker"
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill

        # Control item columns
        items = CONTROL_ITEMS[stmt_type]
        for col_idx, (item_name, is_required) in enumerate(items, start=3):
            col_letter = get_column_letter(col_idx)
            display_name = item_name.replace('_', ' ').title()
            if is_required:
                display_name += " *"
            ws[f'{col_letter}1'] = display_name
            ws[f'{col_letter}1'].font = header_font
            ws[f'{col_letter}1'].fill = header_fill
            ws.column_dimensions[col_letter].width = 18

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 10

        # Data rows
        for row_idx, result in enumerate(results, start=2):
            company = result['company']
            ws[f'A{row_idx}'] = company['company_name']
            ws[f'B{row_idx}'] = company['ticker']

            control_items = result[stmt_type]['control_items']
            error = result[stmt_type]['error']

            if error:
                # If error, mark all cells as failed with error message
                for col_idx, (item_name, _) in enumerate(items, start=3):
                    col_letter = get_column_letter(col_idx)
                    ws[f'{col_letter}{row_idx}'] = "ERROR"
                    ws[f'{col_letter}{row_idx}'].fill = fail_fill
                    ws[f'{col_letter}{row_idx}'].comment = Comment(f"Error: {error}", "System")
            else:
                # Fill in results
                for col_idx, (item_name, _) in enumerate(items, start=3):
                    col_letter = get_column_letter(col_idx)

                    if item_name in control_items:
                        ws[f'{col_letter}{row_idx}'] = "Y"
                        ws[f'{col_letter}{row_idx}'].fill = success_fill
                        ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='center')

                        # Add comment with details
                        line_num = control_items[item_name]
                        plabel = ""
                        if 'line_items' in result[stmt_type]:
                            for item in result[stmt_type]['line_items']:
                                if item.get('stmt_order') == line_num:
                                    plabel = item['plabel']
                                    break

                        comment_text = f"Found at line {line_num}\nPlabel: {plabel}"
                        ws[f'{col_letter}{row_idx}'].comment = Comment(comment_text, "System")
                    else:
                        ws[f'{col_letter}{row_idx}'] = "N"
                        ws[f'{col_letter}{row_idx}'].fill = fail_fill
                        ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='center')
                        ws[f'{col_letter}{row_idx}'].comment = Comment("Not found", "System")

        # Statistics row
        stats_row = len(results) + 3
        ws[f'A{stats_row}'] = "Success Rate"
        ws[f'A{stats_row}'].font = bold_font

        for col_idx, (item_name, _) in enumerate(items, start=3):
            col_letter = get_column_letter(col_idx)
            # Count successes
            success_count = sum(1 for r in results if item_name in r[stmt_type]['control_items'])
            total_count = len(results)
            pct = (success_count / total_count * 100) if total_count > 0 else 0

            ws[f'{col_letter}{stats_row}'] = f"{success_count}/{total_count} ({pct:.0f}%)"
            ws[f'{col_letter}{stats_row}'].font = bold_font

    # =========================================================================
    # SHEET 5: FAILED ITEMS DETAIL
    # =========================================================================
    ws_failed = wb.create_sheet("Failed Items")

    headers = ['Company', 'Ticker', 'CIK', 'Statement', 'Control Item', 'Required', 'Reason']
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws_failed[f'{col_letter}1'] = header
        ws_failed[f'{col_letter}1'].font = header_font
        ws_failed[f'{col_letter}1'].fill = header_fill

    ws_failed.column_dimensions['A'].width = 40
    ws_failed.column_dimensions['B'].width = 10
    ws_failed.column_dimensions['C'].width = 12
    ws_failed.column_dimensions['D'].width = 18
    ws_failed.column_dimensions['E'].width = 40
    ws_failed.column_dimensions['F'].width = 10
    ws_failed.column_dimensions['G'].width = 30

    row = 2
    for result in results:
        company = result['company']

        for stmt_type, stmt_name in [('BS', 'Balance Sheet'), ('IS', 'Income Statement'), ('CF', 'Cash Flow')]:
            control_items = result[stmt_type]['control_items']
            error = result[stmt_type]['error']

            for item_name, is_required in CONTROL_ITEMS[stmt_type]:
                if item_name not in control_items:
                    ws_failed[f'A{row}'] = company['company_name']
                    ws_failed[f'B{row}'] = company['ticker']
                    ws_failed[f'C{row}'] = company['cik']
                    ws_failed[f'D{row}'] = stmt_name
                    ws_failed[f'E{row}'] = item_name
                    ws_failed[f'F{row}'] = "Yes" if is_required else "No"
                    ws_failed[f'G{row}'] = error if error else "Not found in filing"
                    row += 1

    wb.save(output_file)
    print(f"\nExcel report saved: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Batch test financial statement mapping')
    parser.add_argument('--year', type=int, required=True, help='Fiscal year (e.g., 2024)')
    parser.add_argument('--quarter', type=int, required=True, choices=[1,2,3,4], help='Fiscal quarter')
    parser.add_argument('--group', choices=['general', 'financial'], help='Predefined company group')
    parser.add_argument('--form', choices=['10-Q', '10-K'], help='Form type filter')
    parser.add_argument('--ticker', nargs='+', help='Test specific companies by ticker (e.g., MSFT AAPL)')
    parser.add_argument('--cik', nargs='+', help='Test specific companies by CIK (e.g., 789019 320193)')

    args = parser.parse_args()

    # Determine which CIKs to test
    if args.ticker:
        # Lookup CIKs from tickers
        conn = psycopg2.connect(config.get_db_connection())
        cur = conn.cursor(cursor_factory=RealDictCursor)

        test_ciks = []
        for ticker in args.ticker:
            cur.execute("SELECT cik FROM companies WHERE ticker = %s", (ticker.upper(),))
            result = cur.fetchone()
            if result:
                test_ciks.append(result['cik'])
            else:
                print(f"Warning: Ticker not found: {ticker}")

        cur.close()
        conn.close()

        if not test_ciks:
            print("Error: No valid tickers specified")
            return

        group_name = "custom (by ticker)"

    elif args.cik:
        test_ciks = [int(c) for c in args.cik]
        group_name = "custom (by CIK)"

    elif args.group == 'financial':
        test_ciks = FINANCIAL_COMPANIES
        group_name = "financial (25 companies)"

    elif args.group == 'general':
        test_ciks = GENERAL_COMPANIES
        group_name = "general (50 companies)"

    else:
        # Default to general companies
        test_ciks = GENERAL_COMPANIES
        group_name = "general (50 companies, default)"

    print("=" * 80)
    print(f"BATCH MAPPING TEST: {args.year} Q{args.quarter}" + (f" ({args.form})" if args.form else ""))
    print(f"Company Group: {group_name}")
    print("=" * 80)

    # Query database for filings
    conn = psycopg2.connect(config.get_db_connection())
    cur = conn.cursor(cursor_factory=RealDictCursor)

    companies = []
    print(f"\nQuerying filings for {len(test_ciks)} companies...")

    for cik in test_ciks:
        query = """
            SELECT c.company_name, c.ticker, c.cik, c.sic, f.adsh, f.source_year, f.source_quarter, f.form_type
            FROM companies c JOIN filings f ON c.cik = f.cik
            WHERE c.cik = %s::text AND f.source_year = %s AND f.source_quarter = %s
        """
        params = [str(cik), args.year, args.quarter]

        if args.form:
            query += " AND f.form_type = %s"
            params.append(args.form)

        query += " ORDER BY f.filed_date DESC LIMIT 1"

        cur.execute(query, params)
        filing = cur.fetchone()

        if filing:
            # Skip companies without tickers
            if filing['ticker']:
                companies.append(dict(filing))
            else:
                print(f"  Skipping CIK {cik}: no ticker")

    cur.close()
    conn.close()

    print(f"Found {len(companies)}/{len(test_ciks)} companies with filings\n")

    if not companies:
        print("Error: No companies found with filings for this period")
        return

    # Create output folders
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = Path(__file__).parent.parent / 'output' / 'results'
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create statements folder
    statements_folder = output_dir / f'mapping_test_{timestamp}'
    statements_folder.mkdir(exist_ok=True)

    print(f"Financial statements will be saved to: {statements_folder.name}/")
    print(f"Starting batch test...\n")

    # Test each company
    results = []

    for idx, company in enumerate(companies, start=1):
        ticker = company['ticker']
        print(f"[{idx}/{len(companies)}] {ticker:6s} {company['company_name'][:35]:35s}", end=" ... ")

        try:
            year = company['source_year']
            quarter = company['source_quarter']
            cik = company['cik']
            adsh = company['adsh']

            # Test control items
            reconstructor = StatementReconstructor(year=year, quarter=quarter)
            result = test_company(company, reconstructor)
            results.append(result)

            # Quick summary
            bs_found = len(result['BS']['control_items'])
            is_found = len(result['IS']['control_items'])
            cf_found = len(result['CF']['control_items'])

            print(f"BS:{bs_found}/8 IS:{is_found}/8 CF:{cf_found}/6", end=" ... ")

            # Generate financial statements
            try:
                fs_results = map_financial_statements(
                    cik=cik,
                    adsh=adsh,
                    year=year,
                    quarter=quarter,
                    company_name=company['company_name'],
                    ticker=ticker
                )

                # Save to statements folder
                wb = create_excel_workbook(fs_results, company['company_name'], ticker)
                statement_file = statements_folder / f"{ticker}_{cik}_financial_statements.xlsx"
                wb.save(statement_file)
                print(f"Saved")

            except Exception as fs_error:
                print(f"Statement error: {str(fs_error)[:40]}")

        except Exception as e:
            print(f"FAILED: {e}")
            # Still add result with error
            results.append({
                'company': company,
                'BS': {'control_items': {}, 'error': str(e)},
                'IS': {'control_items': {}, 'error': str(e)},
                'CF': {'control_items': {}, 'error': str(e)}
            })

    # Generate test report
    print("\n" + "=" * 80)
    print("Generating test results Excel...")

    output_file = output_dir / f'mapping_test_{timestamp}.xlsx'
    create_excel_report(results, output_file)

    print("=" * 80)
    print("BATCH TEST COMPLETE")
    print(f"Test results: {output_file}")
    print(f"Statements folder: {statements_folder}/ ({len(list(statements_folder.glob('*.xlsx')))} files)")
    print("=" * 80)


if __name__ == "__main__":
    main()
