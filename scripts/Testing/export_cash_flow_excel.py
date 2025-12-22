"""
Cash Flow Statement Excel Exporter
===================================
Exports both reconstructed (original) and standardized cash flow statements to Excel.

Usage:
    python export_cash_flow_excel.py --cik 789019 --adsh 0000950170-24-118967
"""

import sys
import argparse
from pathlib import Path
import yaml
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent / 'src'))

from statement_reconstructor import StatementReconstructor
from pattern_parser import parse_pattern
import psycopg2
from psycopg2.extras import RealDictCursor
from config import config


def load_cash_flow_schema_from_csv():
    """Load cash flow schema from CSV v3"""
    csv_path = Path('docs/Plabel Investigation v3.csv')
    df = pd.read_csv(csv_path)
    cf_df = df[df['Statements'] == 'cash flow statement'].copy()

    schema = {}
    for _, row in cf_df.iterrows():
        target = row['Target']
        pattern = row['Common Variations']
        if pd.notna(target) and pd.notna(pattern):
            schema[target] = pattern

    return schema


def find_control_items(line_items, schema):
    """Find the 3 control items that divide cash flow sections"""
    control_targets = {
        'net cash provided by operating activities': 'operating',
        'net cash provided by investing activities': 'investing',
        'net cash provided by financing activities': 'financing'
    }

    control_lines = {}
    for item in line_items:
        plabel = item['plabel']
        line_num = item.get('stmt_order', 0)

        for target, section in control_targets.items():
            if section in control_lines:
                continue
            pattern = schema.get(target)
            if pattern and parse_pattern(pattern, plabel):
                control_lines[section] = line_num
                break

    return control_lines


def classify_item_section(line_num, control_lines):
    """Classify item into operating/investing/financing based on line position"""
    if not control_lines:
        return 'unknown'

    sorted_controls = sorted(control_lines.items(), key=lambda x: x[1])
    for section, control_line in sorted_controls:
        if line_num <= control_line:
            return section

    return 'supplemental'


def format_currency(value):
    """Format value as currency"""
    if value is None or pd.isna(value):
        return None
    return float(value)


def create_reconstructed_sheet(wb, line_items, periods, company_name, ticker):
    """Create sheet with reconstructed (original) statement"""
    ws = wb.create_sheet("Reconstructed Statement", 0)

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    section_font = Font(bold=True, size=10)
    border = Border(
        bottom=Side(style='thin', color='000000')
    )

    # Title
    ws['A1'] = f"{company_name} ({ticker})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Consolidated Statement of Cash Flows (Reconstructed)"
    ws['A2'].font = Font(size=11)

    # Column headers
    row = 4
    ws[f'A{row}'] = "Line Item"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = border

    for col_idx, period in enumerate(periods, start=2):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{row}'] = period['label']
        ws[f'{col_letter}{row}'].font = header_font
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
        ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='right')

    # Data rows
    row += 1
    current_section = None

    for item in line_items:
        # Check if we're entering a new section (based on control items)
        section = item.get('section', '')
        if section and section != current_section:
            # Add section header
            ws[f'A{row}'] = f"{section.upper()} ACTIVITIES"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1
            current_section = section

        # Line item label
        indent = "  " * item.get('inpth', 0)
        ws[f'A{row}'] = f"{indent}{item['plabel']}"

        # Values for each period
        values = item.get('values', {})
        for col_idx, period in enumerate(periods, start=2):
            col_letter = get_column_letter(col_idx)
            period_key = period['label']
            value = values.get(period_key)

            if value is not None and not pd.isna(value):
                ws[f'{col_letter}{row}'] = format_currency(value)
                ws[f'{col_letter}{row}'].number_format = '#,##0'
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='right')

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 60
    for col_idx in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def create_standardized_sheet(wb, standardized_schema, line_items, periods, company_name, ticker):
    """Create sheet with standardized schema statement"""
    ws = wb.create_sheet("Standardized Statement", 1)

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    section_font = Font(bold=True, size=10)
    border = Border(bottom=Side(style='thin', color='000000'))

    # Title
    ws['A1'] = f"{company_name} ({ticker})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Consolidated Statement of Cash Flows (Standardized Schema)"
    ws['A2'].font = Font(size=11)

    # Column headers
    row = 4
    ws[f'A{row}'] = "Standardized Item"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = border

    for col_idx, period in enumerate(periods, start=2):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{row}'] = period['label']
        ws[f'{col_letter}{row}'].font = header_font
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
        ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='right')

    # Build aggregated values for each period
    # For each target, we need to sum values from all source items across all periods
    target_period_values = {}

    for target, data in standardized_schema.items():
        target_period_values[target] = {}

        # Find all line items that map to this target
        for source_item in data['source_items']:
            # Find the line item with this plabel
            for item in line_items:
                if item['plabel'] == source_item:
                    values = item.get('values', {})
                    for period_label, value in values.items():
                        if period_label not in target_period_values[target]:
                            target_period_values[target][period_label] = 0
                        if value is not None and not pd.isna(value):
                            target_period_values[target][period_label] += value

    # Data rows by section
    row += 1
    sections_order = ['operating', 'financing', 'investing', 'supplemental']

    for section in sections_order:
        section_targets = {k: v for k, v in standardized_schema.items() if v['section'] == section}

        if section_targets:
            # Section header
            ws[f'A{row}'] = f"{section.upper()} ACTIVITIES"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            row += 1

            # Target items
            for target, data in section_targets.items():
                ws[f'A{row}'] = f"  {target}"

                # Values for each period
                period_values = target_period_values.get(target, {})
                for col_idx, period in enumerate(periods, start=2):
                    col_letter = get_column_letter(col_idx)
                    period_label = period['label']
                    value = period_values.get(period_label, 0)

                    if value != 0:
                        ws[f'{col_letter}{row}'] = format_currency(value)
                        ws[f'{col_letter}{row}'].number_format = '#,##0'
                        ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='right')

                row += 1

            row += 1  # Extra space between sections

    # Set column widths
    ws.column_dimensions['A'].width = 60
    for col_idx in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def export_to_excel(cik, adsh, year, quarter, company_name, ticker):
    """Export both reconstructed and standardized statements to Excel"""

    print(f"\n{'='*80}")
    print(f"EXPORTING CASH FLOW STATEMENT TO EXCEL")
    print(f"{'='*80}")
    print(f"\nCompany: {company_name} (Ticker: {ticker}, CIK: {cik})")
    print(f"Filing: {adsh}")
    print(f"Dataset: {year}Q{quarter}")

    # Load schema
    schema = load_cash_flow_schema_from_csv()

    # Reconstruct statement
    print(f"\n📋 Reconstructing statement...")
    reconstructor = StatementReconstructor(year=year, quarter=quarter)
    result = reconstructor.reconstruct_statement_multi_period(
        cik=cik,
        adsh=adsh,
        stmt_type='CF'
    )

    if not result or not result.get('line_items'):
        print("❌ Failed to reconstruct cash flow statement")
        return None

    line_items = result['line_items']
    periods = result.get('periods', [])

    print(f"   ✅ Reconstructed {len(line_items)} line items across {len(periods)} periods")

    # Map items and generate standardized schema
    print(f"\n📋 Mapping to standardized schema...")
    control_lines = find_control_items(line_items, schema)

    mappings = []
    for item in line_items:
        plabel = item['plabel']
        line_num = item.get('stmt_order', 0)
        section = classify_item_section(line_num, control_lines)
        item['section'] = section

        # Try to match against schema patterns
        matched_target = None
        for target, pattern in schema.items():
            if parse_pattern(pattern, plabel):
                matched_target = target
                break

        if matched_target:
            values = item.get('values', {})
            first_value = list(values.values())[0] if values else None

            mappings.append({
                'plabel': plabel,
                'target': matched_target,
                'value': first_value,
                'section': section,
                'confidence': 0.9
            })

    # Aggregate by target
    standardized_schema = {}
    for m in mappings:
        target = m['target']
        if target not in standardized_schema:
            standardized_schema[target] = {
                'section': m['section'],
                'source_items': [],
                'confidence': m['confidence']
            }
        standardized_schema[target]['source_items'].append(m['plabel'])

    print(f"   ✅ Mapped {len(mappings)} items to {len(standardized_schema)} unique targets")

    # Create Excel workbook
    print(f"\n📊 Creating Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create reconstructed statement sheet
    create_reconstructed_sheet(wb, line_items, periods, company_name, ticker)

    # Create standardized statement sheet
    create_standardized_sheet(wb, standardized_schema, line_items, periods, company_name, ticker)

    # Save workbook
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / f"{ticker}_{cik}_cash_flow.xlsx"

    wb.save(output_file)
    print(f"\n✅ Excel file saved to: {output_file}")

    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Export cash flow statement to Excel (reconstructed + standardized)')
    parser.add_argument('--cik', required=True, help='Company CIK')
    parser.add_argument('--adsh', required=True, help='Filing ADSH')

    args = parser.parse_args()

    # Get company info from database
    conn = psycopg2.connect(config.get_db_connection())
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT c.company_name, c.ticker, f.source_year, f.source_quarter
        FROM companies c
        JOIN filings f ON c.cik = f.cik
        WHERE c.cik = %s AND f.adsh = %s
    """, (args.cik, args.adsh))

    info = cur.fetchone()
    cur.close()
    conn.close()

    if not info:
        print(f"❌ Filing not found: CIK {args.cik}, ADSH {args.adsh}")
        sys.exit(1)

    export_to_excel(
        cik=args.cik,
        adsh=args.adsh,
        year=info['source_year'],
        quarter=info['source_quarter'],
        company_name=info['company_name'],
        ticker=info['ticker'] or 'N/A'
    )
