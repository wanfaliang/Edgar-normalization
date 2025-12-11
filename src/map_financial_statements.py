"""
Financial Statements Mapper
===========================
Comprehensive mapper for Balance Sheet, Income Statement, and Cash Flow Statement.

Maps all three statements to standardized schema using direct Python logic (no pattern parsing).

Features:
- Maps BS, IS, CF with consistent naming from CSV v4 "Field Names in DB" column
- Auto-assignment to other_* categories for unmapped items
- 100% coverage for all statements
- ONE YAML file with all three statements
- ONE Excel file with reconstructed + standardized sheets

Usage:
    python map_financial_statements.py --cik 789019 --adsh 0000950170-24-118967
"""

import sys
import argparse
from pathlib import Path
import yaml
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent / 'src'))

from statement_reconstructor import StatementReconstructor
from map_financial_statements_strategy2 import (
    should_use_strategy2,
    map_balance_sheet_strategy2,
    calculate_residuals_strategy2,
    get_balance_sheet_structure_strategy2
)
import psycopg2
from psycopg2.extras import RealDictCursor
from config import config


def normalize(text):
    """Normalize text for matching"""
    if not text:
        return ""
    return text.lower().replace('-', ' ').replace(',', '').replace('  ', ' ').strip()


# ============================================================================
# BALANCE SHEET MAPPING
# ============================================================================

def find_bs_control_items(line_items):
    """
    Find control items for balance sheet.

    Returns dict with line numbers for:
    - total_current_assets (required)
    - total_non_current_assets
    - total_assets (required)
    - total_current_liabilities (required)
    - total_non_current_liabilities
    - total_liabilities
    - total_stockholders_equity (required)
    - total_equity
    - total_liabilities_and_total_equity (required)
    """
    control_lines = {}

    for item in line_items:
        p = normalize(item['plabel'])
        line_num = item.get('stmt_order', 0)

        # 1. total_current_assets (CSV line 10) - REQUIRED
        # Pattern: [contains 'total current assets'] or [contains 'current assets' not contains 'other']
        if 'total_current_assets' not in control_lines:
            if ('total current assets' in p) or ('current assets' in p and 'other' not in p):
                control_lines['total_current_assets'] = line_num

        # 2. total_non_current_assets (CSV line 20)
        # Pattern: [contains 'total' and (contains 'non current assets' or contains 'non-current assets')]
        if 'total_non_current_assets' not in control_lines:
            if 'total' in p and ('non current assets' in p or 'noncurrent assets' in p or 'non-current assets' in p):
                control_lines['total_non_current_assets'] = line_num

        # 3. total_assets (CSV line 21) - REQUIRED
        # Pattern: [equals to 'total assets' or equals to 'assets, total']
        if 'total_assets' not in control_lines:
            if p == 'total assets' or p == 'assets total':
                control_lines['total_assets'] = line_num

        # 4. total_current_liabilities (CSV line 32) - REQUIRED
        # Pattern: [contains 'total current liabilities'] or [contains 'current liabilities' not contains 'other']
        if 'total_current_liabilities' not in control_lines:
            if ('total current liabilities' in p) or ('current liabilities' in p and 'other' not in p):
                control_lines['total_current_liabilities'] = line_num

        # 5. total_non_current_liabilities
        # Pattern: [contains 'total' and (contains 'non current liabilities' or contains 'non-current liabilities' or contains 'long term liabilities')]
        if 'total_non_current_liabilities' not in control_lines:
            if 'total' in p and ('non current liabilit' in p or 'noncurrent liabilit' in p or 'non-current liabilit' in p or 'long term liabilit' in p or 'long-term liabilit' in p):
                control_lines['total_non_current_liabilities'] = line_num

        # 6. total_liabilities (CSV line 41)
        # Pattern: [equals to 'total liabilities' or equals to 'liabilities, total']
        if 'total_liabilities' not in control_lines:
            if p == 'total liabilities' or p == 'liabilities total':
                control_lines['total_liabilities'] = line_num

        # 6. total_stockholders_equity (CSV line 49) - REQUIRED
        # Pattern: [contains 'total' and contains 'equity'] not [contains 'other' or contains 'liabilities' or contains 'total equity']
        if 'total_stockholders_equity' not in control_lines:
            if ('total' in p and 'equity' in p and 'other' not in p and 'liabilit' not in p and p != 'total equity') or ('total stockholders' in p and 'liabliti' not in p) or ('total shareholders' in p and 'liabliti' not in p):
                control_lines['total_stockholders_equity'] = line_num

        # 7. total_equity (CSV line 50)
        # Pattern: [equals to 'total equity' or equals to 'equity, total']
        if 'total_equity' not in control_lines:
            if p == 'total equity' or p == 'equity total':
                control_lines['total_equity'] = line_num

        # 8. total_liabilities_and_total_equity (CSV line 53) - REQUIRED
        # Pattern: [contains 'liabilities' and contains 'equity']
        if 'total_liabilities_and_total_equity' not in control_lines:
            if 'liabilit' in p and ('equity' in p or 'shareholder' in p or 'stockholder' in p):
                control_lines['total_liabilities_and_total_equity'] = line_num

    return control_lines


def classify_bs_section(line_num, control_lines):
    """Classify BS item into section"""
    total_current_assets = control_lines.get('total_current_assets', float('inf'))
    total_assets = control_lines.get('total_assets', float('inf'))
    total_current_liabilities = control_lines.get('total_current_liabilities', float('inf'))
    total_liabilities = control_lines.get('total_liabilities', float('inf'))
    total_stockholders_equity = control_lines.get('total_stockholders_equity', float('inf'))

    if line_num <= total_current_assets:
        return 'current_assets'
    elif line_num <= total_assets:
        return 'non_current_assets'
    elif line_num <= total_current_liabilities:
        return 'current_liabilities'
    elif line_num <= total_liabilities:
        return 'non_current_liabilities'
    elif line_num <= total_stockholders_equity:
        return 'stockholders_equity'
    else:
        return 'equity_total'


def map_bs_item(plabel, line_num, control_lines, tag='', negating=0, datatype=''):
    """Map a balance sheet line item to standardized target"""
    p = normalize(plabel)
    t = tag.lower()  # Lowercase tag for pattern matching
    dt = datatype.lower() if datatype else ''  # Lowercase datatype for matching

    # Get control line numbers
    total_current_assets = control_lines.get('total_current_assets', float('inf'))
    total_non_current_assets = control_lines.get('total_non_current_assets', float('inf'))
    total_assets = control_lines.get('total_assets', float('inf'))
    total_current_liabilities = control_lines.get('total_current_liabilities', float('inf'))
    total_non_current_liabilities = control_lines.get('total_non_current_liabilities', float('inf'))
    total_liabilities = control_lines.get('total_liabilities', float('inf'))
    total_stockholders_equity_line = control_lines.get('total_stockholders_equity', float('inf'))
    total_liabilities_and_equity_line = control_lines.get('total_liabilities_and_total_equity', float('inf'))

    # Map control items DIRECTLY by line number (no duplicate patterns)
    if line_num == total_current_assets:
        return 'total_current_assets'
    if line_num == total_non_current_assets:
        return 'total_non_current_assets'
    if line_num == total_assets:
        return 'total_assets'
    if line_num == total_current_liabilities:
        return 'total_current_liabilities'
    if line_num == total_non_current_liabilities:
        return 'total_non_current_liabilities'
    if line_num == total_liabilities:
        return 'total_liabilities'
    if line_num == total_stockholders_equity_line:
        return 'total_stockholders_equity'
    if line_num == total_liabilities_and_equity_line:
        return 'total_liabilities_and_total_equity'

    # CURRENT ASSETS
    if line_num <= total_current_assets:
        # Cash, cash equivalents and restricted cash (combined) - check FIRST
        if 'cash' in p and 'restricted' in p:
            return 'cash_cash_equivalent_and_restricted_cash'
        # CSV line 4: cash and short-term investments (combined - check FIRST before separates)
        if 'cash and short term' in p:
            return 'cash_and_short_term_investments'
        # CSV line 2: [contains 'cash'] not [contains 'restricted'] not [contains 'cash and short-term investments']
        if ('cash' in p and 'restricted cash' != p) and ('total cash cash equivalents and marketable securities' != p) and ('total cash cash equivalents and short term investments' != p) :
            return 'cash_and_cash_equivalents'
        # CSV line 3: [short-term AND investments] OR [marketable AND securities] OR [marketable AND investments]
        if (( 'investment' in p) or \
           ('marketable' in p and 'securities' in p) or \
           ('marketable' in p and 'investment' in p) or ('investment' in p and 'securit' in p)) and ('total cash cash equivalents and marketable securities' != p) and ('total cash cash equivalents and short term investments' != p):
            return 'short_term_investments'
        # CSV line 5: min{[trade AND receivable] OR [accounts AND receivable] OR [notes AND receivable]}
        if ('trade' in p and ('receivable' in p or 'receivables' in p)) or \
           (('account' in p or 'accounts' in p) and ('receivable' in p or 'receivables' in p)) or \
           (('note' in p or 'notes' in p) and ('receivable' in p or 'receivables' in p)) or ('receivables net' in p and ('other' not in p and 'tax' not in p)):
            return 'account_receivables_net'
        # CSV line 6: other receivables
        if 'other receivable' in p or 'other account receivable' in p or 'other accounts receivable' in p:
            return 'other_receivables'
        # CSV line 7: [inventory OR inventories] OR ([materials and supplies] NOT [inventory])
        if ('inventory' in p or 'inventories' in p) or \
           ('materials and supplies' in p and "inventorynet" in tag):
            return 'inventory'
        if 'prepaid' in p:
            return 'prepaids'
        if 'total' in p and 'current' in p and 'asset' in p:
            return 'total_current_assets'
        # REMOVED: other_current_assets pattern (now calculated as residual)

    # NON-CURRENT ASSETS
    elif line_num <= total_assets:
        if (('property' in p or 'plant' in p or 'equipment' in p or 'ppe' in p or 'fixed assets' in p) and ('net' in p or 'less' in p)) and ('gross' not in tag and 'gross' not in p and 'cost' not in p and 'cost' not in p):
            return 'property_plant_equipment_net'
        if ('investment' in p or 'marketable securities' in p) and line_num > total_current_assets:
            return 'long_term_investments'
        # CSV line 17: goodwill and intangible assets (combined - check FIRST)
        if 'goodwill' in p and ('intangible' in p or 'intangibles' in p):
            return 'goodwill_and_intangible_assets'
        # CSV line 16: [goodwill] NOT [intangible]
        if 'goodwill' in p:
            return 'goodwill'
        # CSV line 15: [intangible OR intangibles] NOT [goodwill]
        if 'intangible' in p or 'intangibles' in p:
            return 'intangible_assets'
        if (('finance' in p or 'capital' in p) and ('lease' in p or 'leases' in p or 'right of use' in p or 'rou' in p)):
            return 'finance_lease_right_of_use_assets'
        if (('operating' in p or 'operation' in p) and ('lease' in p or 'leases' in p or 'right of use' in p or 'rou' in p)):
            return 'operating_lease_right_of_use_assets'
        # Combined lease assets (catch-all after specific operating/finance patterns)
        if 'lease' in p or 'right of use' in p or 'rou' in p:
            return 'lease_assets'
        if 'deferred' in p and 'tax' in p and line_num < total_assets:
            return 'deferred_tax_assets'
        # REMOVED: total_assets pattern - control items mapped by line number only (line 168-169)
        # REMOVED: other_non_current_assets pattern (now calculated as residual)

    # LIABILITIES AND EQUITY (after total_assets)
    elif line_num > total_assets:
        # Special case: CommonStock tags with negating=1 are treasury stock
        if ('commonstock' in t or 'commonshare' in t):
            negating_str = str(negating).strip().upper() if negating else ""
            if negating in (1, True) or negating_str in ("1", "TRUE"):
                return 'treasury_stock'

        # Try EQUITY patterns first (most specific)
        # CSV line 44: [(common stock OR common stocks OR common shares OR common share) AND (cost OR par)]
        # Also match by tag for edge cases - MUST be monetary datatype
        if 'monetary' in dt and (((('common stock' in p or 'common stocks' in p or 'common share' in p or 'common shares' in p) and
             ('cost' in p or 'par' in p or 'issued' in p) and
             ('treasury' not in p and 'purchase' not in p)) or
            (('commonstock' in t or 'commonshare' in t) and ('additional' not in p and 'paid in' not in p and 'excess' not in p and 'surplus' not in p and 'treasury' not in p and 'purchase' not in p))) or ('common stock' == p) or ('common stocks' == p)or ('common shares' == p) or ('common share' == p)):
            return 'common_stock'
        # CSV line 43: [(preferred stock OR preferred stocks) AND (cost OR par)]
        if (('preferred stock' in p or 'preferred stocks' in p) and ('cost' in p or 'par' in p)) or ('preferred stock' == p) or ('preferred stocks' == p):
            return 'preferred_stock'
        # CSV line 46: [(additional OR excess) AND (capital OR proceeds OR fund)]
        if ('additional' in p or 'paid in' in p or 'excess' in p or 'surplus' in p) and \
           ('capital' in p or 'proceeds' in p or 'fund' in p or 'amount' in p)  :
            return 'additional_paid_in_capital'
        # CSV line 45: [(accumulated OR retained) AND (earnings AND deficit)]
        # Note: CSV shows "AND" but likely means "OR" for earnings/deficit
        if ('accumulated' in p or 'retained' in p or 'employed' in p or 'reinvest' in p) and ('earning' in p or 'deficit' in p or 'profit' in p):
            return 'retained_earnings'
        # CSV line 47: [accumulated AND other AND comprehensive]
        # Check tag first - some total_equity items have similar plabels
        if 'accumulated' in p and 'other' in p and 'comprehensive' in p:
            if 'stockholdersequityincludingportionattributabletononcontrolling' in t:
                return 'total_equity'
            return 'accumulated_other_comprehensive_income_loss'
        # CSV line 42: [(treasury stock OR treasury stocks) AND (cost OR par)]
        if (('treasury' in p or 'purchase' in p) and('stock' in p or 'share' in p)) or ('esop' in p or 'option' in p ) :
            return 'treasury_stock'
        # CSV: [contains 'noncontrolling interests in subsidiaries']
        if 'noncontrolling interests in subsidiaries' in p:
            return 'redeemable_non_controlling_interests'
        if 'noncontrolling' in p or 'non controlling' in p or 'minority interest' in p:
            return 'minority_interest'
        if  ('stockholder' in p or 'shareholder' in p or 'owner' in p) and ('equity' in p or 'deficit' in p) and 'liabilit' not in p:
            return 'total_stockholders_equity'
        # REMOVED: other_total_stockholders_equity pattern (now calculated as residual)
        # CSV: [equals to 'total equity' or equals to 'equity, total']
        if p in ['total equity', 'equity total', 'equity, total', 'total deficit']:
            return 'total_equity'
        # REMOVED: total_liabilities_and_total_equity pattern - control items mapped by line number only (line 178-179)

        # Then try LIABILITY patterns
        # Current liabilities
        if line_num <= total_current_liabilities:
            if ('account' in p or 'trade' in p) and ('payable' in p):
                return 'account_payables'
            if 'employ' in p or 'compensation' in p or 'wages' in p or 'salaries' in p or 'payroll' in p:
                return 'accrued_payroll'
            if 'accrued' in p and not any(x in p for x in ['employment', 'compensation', 'wages', 'salaries', 'payroll', 'tax', 'taxes']):
                return 'accrued_expenses'
            # CSV line 30: [unearned OR unexpired] - Check BEFORE short_term_debt to avoid broad "current portion" match
            if ('unearned' in p or 'unexpired' in p) or ('deferred' in p and ('income' in p or 'revenue' in p)) or ('advance' in p):
                return 'deferred_revenue'
            # CSV line 26: [(borrowings OR debt OR notes OR loan OR loans) NOT (long-term)] OR [(one year OR long-term) AND within] OR [current maturities OR current portion]
            if (('borrowing' in p or 'borrowings' in p or 'debt' in p or  'notes' in p or 'loan' in p or 'loans' in p) and
                'long-term' not in p and 'long term' not in p) or \
               ((('one year' in p or 'long-term' in p or 'long term' in p) and 'within' in p) and 'lease' not in p) or \
               (('current maturities' in p or 'current portion' in p or 'current installment' in p) and 'lease' not in p):
                return 'short_term_debt'
            # Dividends payable
            if 'dividend' in p and ('payable' in p or 'liability' in p):
                return 'dividends_payable'
            # CSV line 29: [(payables OR payable) AND income taxes]
            if ('payable' in p or 'accrued' in p or 'liabilit' in p or 'obligation' in p) and 'tax' in p and 'deferred' not in p:
                return 'tax_payables'
            # CSV line 27: [current OR short-term] AND [(finance OR capital) AND (lease OR leases OR right of use OR rou)] AND [position_after # total_assets]
            if \
               ('finance' in p or 'capital' in p) and \
               ('lease' in p or 'leases' in p or 'right of use' in p or 'rou' in p):
                return 'finance_lease_obligations_current'
            if 'operati' in p and 'lease' in p and (line_num > total_assets and line_num <= total_current_liabilities):
                return 'operating_lease_obligations_current'
            # Combined lease obligations current (catch-all after specific operating/finance patterns)
            if 'lease' in p or 'right of use' in p or 'rou' in p:
                return 'lease_obligation_current'
            # REMOVED: other_payables pattern (captured in residual)
            if 'total' in p and 'current' in p and 'liabilit' in p:
                return 'total_current_liabilities'
            # REMOVED: other_current_liabilities pattern (now calculated as residual)

        # Non-current liabilities
        else:
            if (('note payable' in p or 'notes payable' in p) or 'borrowing' in p or 'debt' in p) and line_num > total_current_liabilities:
                return 'long_term_debt'
            # Also capture "term debt" after total_current_liabilities
            if 'term debt' in p or 'long term obligation' in p:
                return 'long_term_debt'
            if ('pension' in p or 'retirement' in p or 'employ' in p) and ('liabilit' in p or 'obligation' in p or 'benefit' in p):
                return 'pension_and_postretirement_benefits'
            if ('deferred revenue' in p or 'unearned' in p) and line_num > total_current_liabilities:
                return 'deferred_revenue_non_current'
            if 'deferred' in p and 'tax' in p and line_num > total_current_liabilities:
                return 'deferred_tax_liabilities_non_current'
            # Tax payables non-current (broader pattern, replaces income_tax_payable_non_current)
            if ('payable' in p or 'accrued' in p or 'liabilit' in p or 'obligation' in p) and 'tax' in p and 'deferred' not in p:
                return 'tax_payables_non_current'
            if ('finance' in p or 'capital' in p) and ('lease' in p or 'leases' in p) and line_num > total_current_liabilities:
                return 'finance_lease_obligations_non_current'
            if ('operating' in p) and ('lease' in p or 'leases' in p) and line_num > total_current_liabilities:
                return 'operating_lease_obligations_non_current'
            # Combined lease obligations non-current (catch-all after specific operating/finance patterns)
            if 'lease' in p or 'right of use' in p or 'rou' in p:
                return 'lease_obligation_non_current'
            if 'commitments' in p or 'contingencies' in p:
                return 'commitments_and_contingencies'
            if p in ['total liabilities', 'liabilities total'] or ('total' in p and 'liabilit' in p and 'current' not in p and 'stockholder' not in p and 'equity' not in p):
                return 'total_liabilities'
            # REMOVED: other_non_current_liabilities pattern (now calculated as residual)

    return None


# ============================================================================
# INCOME STATEMENT MAPPING
# ============================================================================

def find_is_control_items(line_items):
    """
    Find control items for income statement.

    Returns dict with line numbers for:
    - revenue
    - operating_income
    - income_tax_expense (required)
    - net_income (required)
    - eps (required)
    - eps_diluted (required)
    - weighted_average_shares_outstanding
    - weighted_average_shares_outstanding_diluted
    """
    control_lines = {}

    for item in line_items:
        p = normalize(item['plabel'])
        line_num = item.get('stmt_order', 0)
        datatype = item.get('datatype')

        # 1. revenue (CSV line 54)
        # Pattern: [contains 'total revenues' or contains 'net revenue' or contains 'net revenues' or contains 'revenues' or contains 'net sales' or contains 'sales' or contains 'total net sales' or contains 'total net revenue' or contains 'total net sales and revenue'  or (contains 'total' and contains 'revenues')]
        if 'revenue' not in control_lines:
            if ('total revenue' in p or 'net revenue' in p or 'net revenues' in p or 'revenues' in p or \
               'net sales' in p or 'sales' in p or 'total net sales' in p or 'total net revenue' in p or \
               'total net sales and revenue' in p or ('total' in p and 'revenues' in p) or 'revenue'== p or 'revenue:' ==p) and ('marketing' not in p and 'admini' not in p and 'general' not in p):
                if 'cost' not in p and 'deferred' not in p:
                    control_lines['revenue'] = line_num

        # 2. operating_income (CSV line 69)
        # Pattern: [(contains 'operating' or contains 'operation' or contains 'continuing') and (contains 'income' or contains 'loss' or contains 'profit' or contains 'earnings')]
        if 'operating_income' not in control_lines:
            if ('operating' in p or 'operation' in p or 'continuing' in p) and \
               ('income' in p or 'loss' in p or 'profit' in p or 'earnings' in p):
                control_lines['operating_income'] = line_num

        # 3. income_tax_expense (CSV line 72) - REQUIRED
        # Pattern: [(contains 'taxes' or contains 'tax') and (contains 'provision' or contains 'benefit' or contains 'expense' or contains 'expenses')]
        if 'income_tax_expense' not in control_lines:
            if ('taxes' in p or 'tax' in p) and ('provision' in p or 'benefit' in p or 'expense' in p or 'expenses' in p or 'on' in p or 'income' in p) and ('before' not in p):
                control_lines['income_tax_expense'] = line_num

        # 4. net_income (CSV line 75) - REQUIRED
        # Pattern: [(contains 'net income' or contains 'net loss' or contains 'net earings' or contains 'net profit') not (contains 'other' or contains 'continuing' or contains 'discontinuing' or contains 'operating' or contains 'operation')]
        if 'net_income' not in control_lines:
            if ('net income' in p or 'net loss' in p or 'net earings' in p or 'net profit' in p):
                if 'other' not in p and 'continuing' not in p and 'discontinuing' not in p and 'operating' not in p and 'operation' not in p:
                    control_lines['net_income'] = line_num

        # 5. eps (CSV line 78) - REQUIRED
        # Pattern: min{[contains 'basic'] and [datatype = perShare]}
        if 'eps' not in control_lines:
            if 'basic' in p and datatype == 'perShare':
                control_lines['eps'] = line_num

        # 6. eps_diluted (CSV line 79) - REQUIRED
        # Pattern: min{[contains 'diluted'] and [datatype = perShare]}
        if 'eps_diluted' not in control_lines:
            if 'diluted' in p and datatype == 'perShare':
                control_lines['eps_diluted'] = line_num

        # 7. weighted_average_shares_outstanding (CSV line 80)
        # Pattern: [contains 'basic'] and [datatype = shares]
        if 'weighted_average_shares_outstanding' not in control_lines:
            if 'basic' in p and datatype == 'shares':
                control_lines['weighted_average_shares_outstanding'] = line_num

        # 8. weighted_average_shares_outstanding_diluted (CSV line 81)
        # Pattern: [contains 'diluted'] and [datatype = shares]
        if 'weighted_average_shares_outstanding_diluted' not in control_lines:
            if 'diluted' in p and datatype == 'shares':
                control_lines['weighted_average_shares_outstanding_diluted'] = line_num

    return control_lines


def classify_is_section(line_num, control_lines):
    """Classify IS item into section"""
    revenue = control_lines.get('revenue', 0)
    gross_profit = control_lines.get('gross_profit', float('inf'))
    operating_income = control_lines.get('operating_income', float('inf'))
    income_before_tax = control_lines.get('income_before_tax', float('inf'))
    net_income = control_lines.get('net_income', float('inf'))

    if line_num <= gross_profit:
        return 'revenue_and_cogs'
    elif line_num <= operating_income:
        return 'operating_expenses'
    elif line_num <= income_before_tax:
        return 'non_operating'
    elif line_num <= net_income:
        return 'tax_and_net_income'
    else:
        return 'below_net_income'


def map_is_item(plabel, line_num, control_lines, datatype=None):
    """Map an income statement line item to standardized target"""
    p = normalize(plabel)

    # Map control items DIRECTLY by line number (no duplicate patterns)
    if line_num == control_lines.get('revenue'):
        return 'revenue'
    if line_num == control_lines.get('gross_profit'):
        return 'gross_profit'
    if line_num == control_lines.get('operating_income'):
        return 'operating_income'
    if line_num == control_lines.get('income_before_tax'):
        return 'income_before_tax'
    if line_num == control_lines.get('income_tax_expense'):
        return 'income_tax_expense'
    if line_num == control_lines.get('net_income'):
        return 'net_income'
    if line_num == control_lines.get('eps'):
        return 'eps'
    if line_num == control_lines.get('eps_diluted'):
        return 'eps_diluted'
    if line_num == control_lines.get('weighted_average_shares_outstanding'):
        return 'weighted_average_shares_outstanding'
    if line_num == control_lines.get('weighted_average_shares_outstanding_diluted'):
        return 'weighted_average_shares_outstanding_diluted'

    # Revenue - CSV line 54
    if 'revenue' in p or 'sales' in p:
        if 'marketing' not in p and 'admini' not in p and 'general' not in p:
            if 'cost' not in p and 'deferred' not in p and 'unearned' not in p:
                return 'revenue'

    # Cost of revenue - CSV line 55
    if ('cost' in p and ('revenue' in p or 'sales' in p or 'sold' in p or 'sale' in p)) or 'cogs' in p:
        return 'cost_of_revenue'

    # Gross profit - CSV line 56
    if 'gross profit' in p or 'gross margin' in p or 'gross income' in p:
        return 'gross_profit'

    # CSV line 63: Cost and expenses
    if 'total costs and expenses' in p or 'total cost and expenses' in p or 'total operating costs and expenses' in p or 'total operating costs' in p:
        return 'cost_and_expenses'

    section = classify_is_section(line_num, control_lines)

    # Operating expenses - match CSV field names
    # CSV line 57: R&D
    if 'research' in p or 'development' in p or 'r&d' in p or 'technology' in p:
        return 'research_and_development_expenses'

    # CSV line 60: SG&A (combined)
    if ('sales' in p or 'marketing' in p or 'selling' in p or 'sale' in p or 'advertising' in p or 'promotion' in p) and 'administrative' in p:
        return 'selling_general_and_administrative_expenses'

    # CSV line 59: Sales and marketing (not administrative)
    if ('sales' in p or 'marketing' in p or 'selling' in p or 'sale' in p or 'advertising' in p or 'promotion' in p) and 'administrative' not in p:
        return 'sales_and_marketing_expenses'

    # CSV line 58: G&A (not selling/marketing)
    if ('general' in p and 'administrative' in p) and 'selling' not in p and 'marketing' not in p and 'advertising' not in p and 'promotion' not in p and 'sales' not in p:
        return 'general_and_administrative_expenses'

    # CSV line 67: Depreciation and amortization (IS version)
    if 'depreciation' in p or 'amortization' in p:
        return 'depreciation_and_amortization'

    # CSV line 62: Operating expenses
    if 'total operating expenses' in p or 'total expenses' in p:
        return 'operating_expenses'

    # CSV line 61: Other expenses
    if 'other' in p and ('expense' in p or 'expenses' in p or 'income' in p):
        return 'other_expenses'

    # Operating income - CSV line 69
    if ('operating' in p or 'operation' in p or 'continuing' in p) and ('income' in p or 'loss' in p or 'profit' in p or 'earnings' in p):
        return 'operating_income'

    # CSV line 64: Interest income
    if 'interest' in p and 'income' in p:
        return 'interest_income'

    # CSV line 65: Interest expense
    if ('interest' in p or 'interests' in p) and ('expense' in p or 'expenses' in p):
        return 'interest_expense'

    # CSV line 66: Net interest income
    if ('interest' in p or 'interests' in p) and ('expense' in p or 'expenses' in p) and 'income' in p:
        return 'net_interest_income'

    # CSV line 68: Total other income/expenses net
    if 'other' in p and ('income' in p or 'gains' in p or 'gain' in p or 'loss' in p or 'losses' in p or 'expenses' in p):
        return 'total_other_income_expenses_net'

    # CSV line 70: Non-operating income
    if ('non-operating' in p or 'non-operation' in p or 'discontinued' in p) and ('income' in p or 'loss' in p or 'earnings' in p or 'expenses' in p or 'expense' in p):
        return 'non_operating_income'

    # Income before tax - CSV line 71
    if 'pretax' in p or ('before' in p and ('tax' in p or 'taxes' in p)):
        return 'income_before_tax'

    # Income tax expense - CSV line 72
    if ('tax' in p or 'taxes' in p) and ('provision' in p or 'benefit' in p or 'expense' in p or 'expenses' in p) and ('before' not in p):
        return 'income_tax_expense'

    # CSV line 73: Net income from continuing operations
    if ('net income' in p or 'net loss' in p or 'net earnings' in p) and 'continuing' in p:
        return 'net_income_from_continuing_operations'

    # CSV line 74: Net income from discontinued operations
    if ('net income' in p or 'net loss' in p or 'net earnings' in p or 'income' in p or 'loss' in p or 'earnings' in p or 'deficit' in p) and 'discontinued' in p:
        return 'net_income_from_discontinued_operations'

    # CSV line 76: Net income attributed to non-controlling interests
    if ('net income' in p or 'net loss' in p or 'net earnings' in p or 'net profit' in p) and 'attributable to' in p and ('non-controlling' in p or 'noncontrolling' in p or 'minority' in p):
        return 'net_income_attributed_to_non_controlling_interests'

    # CSV line 77: Net income attributable to controlling interests
    if ('net income' in p or 'net loss' in p or 'net earnings' in p or 'net profit' in p) and 'attributable to' in p and 'non-controlling' not in p and 'noncontrolling' not in p and 'minority' not in p:
        return 'net_income_attributable_to_controlling_interests'

    # CSV line 75: Net income (general)
    if ('net income' in p or 'net loss' in p or 'net earnings' in p or 'net profit' in p) and 'other' not in p and 'continuing' not in p and 'discontinuing' not in p and 'operating' not in p and 'operation' not in p:
        return 'net_income'

    # EPS and Shares - use datatype metadata to distinguish - match CSV field names
    if section == 'below_net_income':
        # Use datatype to distinguish EPS (perShare) from share counts (shares)
        if p == 'basic' or 'earnings per share' in p or 'eps' in p:
            if datatype == 'perShare' or 'earnings per share' in p or 'eps' in p:
                return 'eps'
            elif datatype == 'shares' or 'weighted average' in p:
                return 'weighted_average_shares_outstanding'
        if p == 'diluted':
            if datatype == 'perShare':
                return 'eps_diluted'
            elif datatype == 'shares':
                return 'weighted_average_shares_outstanding_diluted'
        # Explicit matches
        if 'weighted average' in p and 'shares' in p:
            if 'diluted' in p:
                return 'weighted_average_shares_outstanding_diluted'
            else:
                return 'weighted_average_shares_outstanding'

    return None


# ============================================================================
# CASH FLOW MAPPING
# ============================================================================

def find_cf_control_items(line_items):
    """
    Find control items for cash flow.

    Returns dict with line numbers for:
    - net_income (required)
    - net_cash_provided_by_operating_activities (required)
    - net_cash_provided_by_investing_activities (required)
    - net_cash_provided_by_financing_activities (required)
    - cash_at_beginning_of_period (required)
    - cash_at_end_of_period (required)
    """
    control_lines = {}

    for item in line_items:
        p = normalize(item['plabel'])
        line_num = item.get('stmt_order', 0)
        iord = item.get('iord', '')
        ddate = item.get('ddate', '')
        qtrs = item.get('qtrs', '')
        tag = normalize(item.get('tag', ''))

        # 1. net_income (CSV line 93) - REQUIRED
        # Pattern: [ contains 'net income' or 'net earnings' or 'net income (loss)']
        if 'net_income' not in control_lines:
            if 'net income' in p or 'net earnings' in p or 'net income (loss)' in p or 'profit' in p or 'net loss' in p or ('net' in p and 'loss' in p):
                control_lines['net_income'] = line_num

        # 2. net_cash_provided_by_operating_activities (CSV line 117) - REQUIRED
        # Pattern: [contains 'cash' and (contains 'operating' or contains 'operations')] not [contains 'other' or contains 'others']
        if 'net_cash_provided_by_operating_activities' not in control_lines:
            if ('cash' in p and ('operating' in p or 'operations' in p) and 'other' not in p and 'others' not in p) or ('total operating' in p):
                control_lines['net_cash_provided_by_operating_activities'] = line_num

        # 3. net_cash_provided_by_investing_activities (CSV line 126) - REQUIRED
        # Pattern: [contains 'cash' and contains 'investing'] not [contains 'other' or contains 'others']
        if 'net_cash_provided_by_investing_activities' not in control_lines:
            if ('cash' in p and 'investing' in p and 'other' not in p and 'others' not in p) or ('total investing' in p):
                control_lines['net_cash_provided_by_investing_activities'] = line_num

        # 4. net_cash_provided_by_financing_activities (CSV line 152) - REQUIRED
        # Pattern: [contains 'cash' and contains 'financing'] not [contains 'other' or  contains 'others']
        if 'net_cash_provided_by_financing_activities' not in control_lines:
            if ('cash' in p and 'financing' in p and 'other' not in p and 'others' not in p) or ('total financing' in p):
                control_lines['net_cash_provided_by_financing_activities'] = line_num

        # 5. cash_at_beginning_of_period (CSV line 156) - REQUIRED
        # Pattern: [iord='I' (instant) and (contains 'cash' in plabel OR tag) and contains 'beginning']
        # Use TAG metadata (iord) like reconstructor does
        # Handle cases like Starbucks where plabel is just "Beginning of period" without 'cash'
        if 'cash_at_beginning_of_period' not in control_lines:
            if iord == 'I':
                # Check if it's a cash-related item (plabel or tag contains 'cash')
                is_cash_item = 'cash' in p or 'cash' in tag
                if is_cash_item:
                    if 'beginning' in p or 'begin' in p or qtrs == '0':
                        control_lines['cash_at_beginning_of_period'] = line_num

        # 6. cash_at_end_of_period (CSV line 155) - REQUIRED
        # Pattern: [iord='I' (instant) and (contains 'cash' in plabel OR tag) and contains 'end']
        # Use TAG metadata (iord) like reconstructor does
        # Handle cases like Starbucks where plabel is just "End of period" without 'cash'
        if 'cash_at_end_of_period' not in control_lines:
            if iord == 'I':
                # Check if it's a cash-related item (plabel or tag contains 'cash')
                is_cash_item = 'cash' in p or 'cash' in tag
                if is_cash_item:
                    if 'end' in p or 'ending' in p or 'close' in p or qtrs == '0':
                        # Prefer ending over beginning (ending usually appears later)
                        if 'cash_at_beginning_of_period' in control_lines:
                            control_lines['cash_at_end_of_period'] = line_num
                        elif 'beginning' not in p and 'begin' not in p:
                            control_lines['cash_at_end_of_period'] = line_num

    return control_lines


def classify_cf_section(line_num, control_lines):
    """Classify CF item into section - returns 'operating', 'investing', 'financing', or 'supplemental'"""
    if not control_lines:
        return 'unknown'

    operating_end = control_lines.get('net_cash_provided_by_operating_activities', 0)
    investing_end = control_lines.get('net_cash_provided_by_investing_activities', 0)
    financing_end = control_lines.get('net_cash_provided_by_financing_activities', 0)

    # Classify based on which section this line falls into
    if line_num <= operating_end:
        return 'operating'
    elif line_num <= investing_end:
        return 'investing'
    elif line_num <= financing_end:
        return 'financing'
    else:
        return 'supplemental'


def map_cf_item(plabel, line_num, control_lines, tag='', line_items=None):
    """Map a cash flow line item to standardized target - matches structure field names"""
    p = normalize(plabel)
    t = normalize(tag)

    # Find position of working capital items for other_adjustments check
    wc_positions = []
    if line_items:
        for item in line_items:
            item_p = normalize(item['plabel'])
            item_line = item.get('stmt_order', float('inf'))
            # Check for accounts_receivables, inventory, or accounts_payables
            if (('account' in item_p or 'accounts' in item_p) and ('receivable' in item_p or 'receivables' in item_p)) or \
               ('inventory' in item_p or 'inventories' in item_p) or \
               (('account' in item_p or 'accounts' in item_p) and ('payable' in item_p or 'payables' in item_p)):
                wc_positions.append(item_line)
    min_wc_position = min(wc_positions) if wc_positions else float('inf')

    # Map control items DIRECTLY by line number (no duplicate patterns)
    if line_num == control_lines.get('net_income'):
        return 'net_income_starting_line'
    if line_num == control_lines.get('net_cash_provided_by_operating_activities'):
        return 'net_cash_provided_by_operating_activities'
    if line_num == control_lines.get('net_cash_provided_by_investing_activities'):
        return 'net_cash_provided_by_investing_activities'
    if line_num == control_lines.get('net_cash_provided_by_financing_activities'):
        return 'net_cash_provided_by_financing_activities'
    if line_num == control_lines.get('cash_at_beginning_of_period'):
        return 'cash_at_beginning_of_period'
    if line_num == control_lines.get('cash_at_end_of_period'):
        return 'cash_at_end_of_period'

    # Apply ALL specific patterns WITHOUT section restrictions

    # ========================================================================
    # OPERATING ACTIVITIES
    # ========================================================================

    # CSV line 94: Depreciation and amortization (combined)
    if ('depreciation' in p) and ('amortization' in p):
        return 'depreciation_and_amortization'

    # CSV line 95: Depreciation only
    if 'depreciation' in p and 'amortization' not in p:
        return 'depreciation'

    # CSV line 96: Amortization only
    if 'amortization' in p and 'depreciation' not in p:
        return 'amortization'

    # CSV line 98: Impairments
    if 'impairment' in p or 'impairments' in p:
        return 'impairments'

    # CSV line 99: Pension and postretirement
    if 'pension' in p or 'postretirement' in p:
        return 'pension_and_postretirement'

    # CSV line 100: Stock-based compensation
    if ('stock-based' in p or 'stock based' in p or 'share-based' in p or 'share based' in p) and 'compensation' in p:
        return 'stock_based_compensation'

    # CSV line 101: Non-operating expense/income
    if ('non-operating' in p or 'non operating' in p) and ('expense' in p or 'expenses' in p or 'income' in p):
        return 'non_operating_expense_income'
    if 'non-cash' in p and 'expenses' in p:
        return 'non_operating_expense_income'

    # CSV line 102: Investment gains/losses
    if ('gain' in p or 'gains' in p or 'loss' in p or 'losses' in p) and ('investment' in p or 'investments' in p):
        return 'investment_gains_losses'

    # CSV line 97: Deferred income tax
    if 'deferred' in p and ('tax' in p or 'taxes' in p):
        return 'deferred_income_tax'

    # CSV: Other adjustments - exact match "other" or "other, net" AND position_before working capital items
    if p in ['other', 'other net', 'other, net'] and line_num < min_wc_position:
        return 'other_adjustments'

    # CSV line 104: Change in other working capital
    if 'other working capital' in p or ('other' in p and 'assets' in p):
        return 'change_in_other_working_capital'

    # CSV line 105: Accounts receivables
    if ('account' in p or 'accounts' in p) and ('receivable' in p or 'receivables' in p):
        return 'accounts_receivables'

    # CSV line 106: Vendor receivables
    if 'vendor' in p and ('receivable' in p or 'receivables' in p):
        return 'vendor_receivables'

    # CSV line 107: Inventory
    if 'inventory' in p or 'inventories' in p:
        return 'inventory'

    # CSV line 108: Prepaids
    if 'prepaid' in p or 'prepaids' in p or 'prepayment' in p or 'prepayments' in p:
        return 'prepaids'

    # CSV line 109: Accounts payables
    if ('account' in p or 'accounts' in p) and ('payable' in p or 'payables' in p):
        return 'accounts_payables'

    # CSV line 110: Accrued expenses
    if 'accrued' in p:
        return 'accrued_expenses'

    # CSV line 111: Unearned revenue
    if 'unearned' in p:
        return 'unearned_revenue'

    # CSV line 112: Income taxes payable
    if 'income taxes' in p and ('payable' in p or 'payables' in p):
        return 'income_taxes_payable'

    # CSV line 113: Income taxes (not payable, not deferred, not paid)
    if ('income taxes' in p or 'income tax' in p) and 'payable' not in p and 'deferred' not in p and 'paid' not in p:
        return 'income_taxes'

    # CSV line 114: Other assets
    if 'other' in p and 'assets' in p:
        return 'other_assets'

    # CSV line 115: Other liabilities
    if 'other' in p and 'liabilities' in p:
        return 'other_liabilities'

    # ========================================================================
    # INVESTING ACTIVITIES
    # ========================================================================

    # CSV line 118: Capital expenditures (investments in PP&E)
    if ('expenditure' in p or 'expenditures' in p or 'purchase' in p or 'purchases' in p or 'acquisition' in p or 'acquisitions' in p or 'addition' in p or 'additions' in p) and \
       ('property' in p or 'plant' in p or 'plants' in p or 'equipment' in p or 'equipments' in p or 'ppe' in p):
        return 'investments_in_property_plant_and_equipment'

    # CSV line 119: Proceeds from sales of PP&E
    if ('proceeds' in p or 'sale' in p or 'sales' in p or 'disposition' in p) and \
       ('property' in p or 'plant' in p or 'plants' in p or 'equipment' in p or 'equipments' in p or 'ppe' in p):
        return 'proceeds_from_sales_of_ppe'

    # CSV line 120: Acquisitions of business
    if ('business' in p or 'businesses' in p) and ('acquisition' in p or 'acquisitions' in p):
        return 'acquisitions_of_business_net'

    # CSV line 121: Proceeds from divestiture
    if ('divestiture' in p or 'sale' in p or 'sales' in p or 'disposition' in p) and ('business' in p or 'businesses' in p):
        return 'proceeds_from_divestiture'

    # CSV line 122: Other acquisitions and investments
    if 'other' in p and ('acquisition' in p or 'acquisitions' in p or 'investments' in p):
        return 'other_aquisitons_and_investments'

    # CSV line 123: Purchases of investments
    if ('purchase' in p or 'purchases' in p or 'acquisition' in p or 'acquisitions' in p) and ('investment' in p or 'investments' in p or 'securities' in p):
        return 'purchases_of_investments'

    # CSV line 124: Sales/maturities of investments
    if ('sale' in p or 'sales' in p or 'maturity' in p or 'maturities' in p or 'proceeds' in p) and ('investment' in p or 'investments' in p or 'securities' in p):
        return 'sales_maturities_of_investments'

    # ========================================================================
    # FINANCING ACTIVITIES
    # ========================================================================

    # CSV line 127: Short-term debt issuance
    if 'short-term' in p and ('issuance' in p or 'proceeds' in p):
        return 'short_term_debt_issuance'

    # CSV line 128: Long-term debt issuance
    if 'long-term' in p and ('issuance' in p or 'proceeds' in p):
        return 'long_term_net_debt_issuance'

    # CSV line 129: Short-term debt repayment
    if 'short-term' in p and ('repayment' in p or 'repayments' in p):
        return 'short_term_debt_repayment'

    # CSV line 130: Long-term debt repayment
    if 'long-term' in p and ('repayment' in p or 'repayments' in p):
        return 'long_term_net_debt_repayment'

    # CSV line 131: Change in short-term debt net
    if ('change' in p or 'changes' in p) and ('short-term debt' in p or 'short-term' in p):
        return 'change_in_short_term_debt_net'

    # CSV line 132: Commercial paper net
    if 'commercial paper' in p:
        return 'commercial_paper_net'

    # CSV line 133: Change in long-term debt net
    if ('change' in p or 'changes' in p) and ('long-term debt' in p or 'long-term' in p):
        return 'change_in_long_term_debt_net'

    # CSV line 134: Term debt issuance (not short/long term specific)
    if 'debt' in p and ('issuance' in p or 'proceeds' in p) and 'short-term' not in p and 'long-term' not in p:
        return 'term_debt_issuance'

    # CSV line 135: Term debt repayment (not short/long term specific)
    if 'debt' in p and ('repayment' in p or 'repayments' in p) and 'short-term' not in p and 'long-term' not in p:
        return 'term_debt_repayment'

    # CSV line 136: Change in term debt
    if 'debt' in p and ('change' in p or 'changes' in p) and 'short-term' not in p and 'long-term' not in p:
        return 'change_in_term_debt'

    # CSV line 137: Finance lease repayment
    if ('finance lease' in p or 'finance leases' in p) and ('repayment' in p or 'repayments' in p or 'principal' in p):
        return 'finance_lease_repayment'

    # CSV line 138: Financing obligations repayment
    if ('financing obligation' in p or 'financing obligations' in p) and ('repayment' in p or 'repayments' in p or 'principal' in p):
        return 'financing_obligations_repayment'

    # CSV line 139: Net stock issuance
    if 'net stock issuance' in p:
        return 'net_stock_issuance'

    # CSV line 140: Net common stock issuance
    if 'net' in p and ('common stock' in p or 'common stocks' in p) and ('issuance' in p or 'proceeds' in p):
        return 'net_common_stock_issuance'

    # CSV line 141: Common stock issuance (not net)
    if ('common stock' in p or 'common stocks' in p or 'shares' in p) and ('issuance' in p or 'issued' in p or 'proceeds' in p) and 'net' not in p:
        return 'common_stock_issuance'

    # CSV line 142: Taxes on share settlement
    if ('tax' in p or 'taxes' in p) and ('share' in p or 'shares' in p) and ('settlement' in p or 'vesting' in p):
        return 'taxes_on_share_settlement'

    # CSV line 143: Net preferred stock issuance
    if 'net' in p and ('preferred stock' in p or 'preferred stocks' in p) and ('issuance' in p or 'proceeds' in p):
        return 'net_preferred_stock_issuance'

    # CSV line 144: Preferred stock issuance (not net)
    if ('preferred stock' in p or 'preferred stocks' in p) and ('issuance' in p or 'proceeds' in p) and 'net' not in p:
        return 'preferred_stock_issuance'

    # CSV line 145: Common stock repurchased
    if (('treasury' in p or 'share' in p or 'stock' in p or 'stocks' in p) and ('purchase' in p or 'purchases' in p)) or \
       ('repurchase' in p or 'repurchases' in p or 'repurchased' in p):
        return 'common_stock_repurchased'

    # CSV line 146: Proceeds from issuance of stock (special purpose)
    if (('treasury stock' in p or 'treasury stocks' in p) and ('issuance' in p or 'proceeds' in p)) or \
       ('proceeds' in p and 'options' in p) or ('proceeds' in p and 'compensation' in p):
        return 'proceeds_from_issuance_of_stock_sp'

    # CSV line 147-149: Dividends paid
    if 'dividend' in p or 'dividends' in p:
        if 'common' in p and 'preferred' not in p:
            return 'common_dividends_paid'
        elif 'preferred' in p and 'common' not in p:
            return 'preferred_dividends_paid'
        else:
            return 'dividends_paid'

    # CSV line 150: Issuance costs
    if 'issuance costs' in p:
        return 'issuance_costs'

    # ========================================================================
    # OTHER ITEMS
    # ========================================================================

    # CSV line 153: Effect of foreign exchange rate changes
    if 'exchange' in p or 'foreign currency' in p or 'foreign currencies' in p:
        return 'effect_of_foreign_exchanges_rate_changes_on_cash'

    # CSV line 154: Net change in cash
    if ('net change' in p or 'net increase' in p or 'net decrease' in p or 'increase' in p or 'decrease' in p) and 'cash' in p and \
       'operating' not in p and 'investing' not in p and 'financing' not in p:
        return 'net_change_in_cash'

    # CSV line 160: Income taxes paid
    if 'income taxes' in p and ('payment' in p or 'payments' in p or 'paid' in p):
        return 'income_taxes_paid'

    # CSV line 161: Interest paid
    if ('payment' in p or 'payments' in p or 'paid' in p) and ('interest' in p or 'interests' in p):
        return 'interest_paid'

    # Section totals
    if 'net cash' in p:
        if 'operating' in p or 'operations' in p:
            return 'net_cash_provided_by_operating_activities'
        elif 'investing' in p:
            return 'net_cash_provided_by_investing_activities'
        elif 'financing' in p:
            return 'net_cash_provided_by_financing_activities'

    # Supplemental - cash beginning/ending
    is_cash_item = 'cash' in p or 'cash' in t
    if is_cash_item and ('beginning' in p or 'start' in p):
        return 'cash_at_beginning_of_period'
    if is_cash_item and ('end' in p or 'close' in p or 'ending' in p):
        return 'cash_at_end_of_period'

    # CSV line 116, 125, 151: ONLY use section for "other" items (no distinctive keywords)
    if 'other' in p:
        section = classify_cf_section(line_num, control_lines)
        if section == 'operating':
            return 'other_operating_activities'
        elif section == 'investing':
            return 'other_investing_activities'
        elif section == 'financing':
            return 'other_financing_activities'

    return None


# ============================================================================
# MAIN MAPPING ORCHESTRATION
# ============================================================================

def map_statement(stmt_type, line_items, control_lines):
    """Map a statement's line items to standardized targets"""
    mappings = []
    target_to_plabels = defaultdict(list)

    for item in line_items:
        plabel = item['plabel']
        line_num = item.get('stmt_order', 0)

        # Get value for first period
        values = item.get('values', {})
        if isinstance(values, dict) and len(values) > 0:
            value = list(values.values())[0]
        else:
            value = None

        # Map based on statement type
        if stmt_type == 'BS':
            section = classify_bs_section(line_num, control_lines)
            tag = item.get('tag', '')
            negating = item.get('negating', item.get('NEGATING', 0))
            datatype = item.get('datatype', '')
            target = map_bs_item(plabel, line_num, control_lines, tag, negating, datatype)
            # REMOVED: auto-assignment logic (other_* now calculated as residuals)

        elif stmt_type == 'IS':
            section = classify_is_section(line_num, control_lines)
            datatype = item.get('datatype')
            target = map_is_item(plabel, line_num, control_lines, datatype)

        elif stmt_type == 'CF':
            section = classify_cf_section(line_num, control_lines)
            tag = item.get('tag', '')
            target = map_cf_item(plabel, line_num, control_lines, tag, line_items)

            # Auto-assign unmapped items to other_*_activities
            if not target:
                if section == 'operating':
                    target = 'other_operating_activities'
                elif section == 'investing':
                    target = 'other_investing_activities'
                elif section == 'financing':
                    target = 'other_financing_activities'

        else:
            section = 'unknown'
            target = None

        if target:
            mappings.append({
                'plabel': plabel,
                'target': target,
                'value': value,
                'section': section
            })
            target_to_plabels[target].append((plabel, line_num))

    # Special handling for total_stockholders_equity: if multiple items match,
    # the one with smaller line number is total_stockholders_equity,
    # the one(s) with greater line number(s) are total_equity
    if stmt_type == 'BS' and 'total_stockholders_equity' in target_to_plabels:
        items = target_to_plabels['total_stockholders_equity']
        if len(items) >= 2:
            # Sort by line number
            items_sorted = sorted(items, key=lambda x: x[1])
            # Keep first (smallest line_num) as total_stockholders_equity
            target_to_plabels['total_stockholders_equity'] = [items_sorted[0]]
            # Move rest to total_equity
            if 'total_equity' not in target_to_plabels:
                target_to_plabels['total_equity'] = []
            target_to_plabels['total_equity'].extend(items_sorted[1:])
            # Update mappings list
            for i, mapping in enumerate(mappings):
                if mapping['target'] == 'total_stockholders_equity':
                    plabel_match = mapping['plabel']
                    # Find line_num for this plabel in items_sorted
                    for plabel, line_num in items_sorted[1:]:
                        if plabel == plabel_match:
                            mappings[i]['target'] = 'total_equity'
                            break

    return mappings, target_to_plabels


def aggregate_by_target(target_to_plabels, line_items):
    """Aggregate values by target across all source items and all periods"""
    standardized = {}

    # Get all unique period labels from line items
    all_periods = set()
    for item in line_items:
        values = item.get('values', {})
        if isinstance(values, dict):
            all_periods.update(values.keys())

    for target, source_items in target_to_plabels.items():
        # Calculate aggregated values for EACH period
        period_values = {}
        for period_label in all_periods:
            period_total = 0
            for plabel, line_num in source_items:
                for item in line_items:
                    if item['plabel'] == plabel and item.get('stmt_order', 0) == line_num:
                        values = item.get('values', {})
                        if isinstance(values, dict):
                            value = values.get(period_label)
                            if value and not pd.isna(value):
                                period_total += value
                        break
            if period_total != 0:
                period_values[period_label] = period_total

        # For backwards compatibility, also store first period as total_value
        total_value = list(period_values.values())[0] if period_values else None

        standardized[target] = {
            'total_value': total_value,
            'period_values': period_values,
            'count': len(source_items),
            'source_items': [plabel for plabel, _ in source_items]
        }

    return standardized


def validate_and_calculate_bs_residuals(standardized, control_lines, sic_code=None):
    """
    Validate balance sheet control totals and calculate residual 'other_*' items.

    Strategic approach:
    Step 0: Check if financial sector (SIC 6000-6999) - skip for now
    Step 1: Validate core totals (total_assets == total_liabilities_and_total_equity)
    Step 2: Determine total_liabilities
    Step 3: Calculate current/non-current splits
    Step 4: Calculate residual other_* items

    Returns:
        dict: Updated standardized dict with residuals, or None if validation fails
        str: Status message ('success', 'failed: reason')
    """

    # Step 0: Check if financial sector
    if sic_code and isinstance(sic_code, (int, str)):
        try:
            sic_int = int(sic_code)
            if 6000 <= sic_int <= 6999:
                return standardized, 'skipped: financial sector'
        except (ValueError, TypeError):
            pass

    # Helper function: check if values are equal within tolerance
    def is_equal_within_tolerance(val1, val2, tolerance=0.001):
        """Check if two values are equal within 0.1% tolerance"""
        if val1 is None or val2 is None:
            return False
        try:
            v1 = float(val1)
            v2 = float(val2)
            if v1 == 0 and v2 == 0:
                return True
            avg = (abs(v1) + abs(v2)) / 2
            if avg == 0:
                return v1 == v2
            diff_pct = abs(v1 - v2) / avg
            return diff_pct < tolerance
        except (ValueError, TypeError):
            return False

    # Helper function: get value for first period
    def get_value(field_name):
        """Get first period value from standardized dict"""
        if field_name not in standardized:
            return None
        period_values = standardized[field_name].get('period_values', {})
        if not period_values:
            return None
        return list(period_values.values())[0]

    # Helper function: get all period values
    def get_period_values(field_name):
        """Get all period values from standardized dict"""
        if field_name not in standardized:
            return {}
        return standardized[field_name].get('period_values', {})

    # Helper function: set value with all periods
    def set_value(field_name, period_values):
        """Set standardized value with all periods"""
        if not period_values:
            return
        standardized[field_name] = {
            'total_value': list(period_values.values())[0],
            'period_values': period_values,
            'count': 1,
            'source_items': ['[calculated residual]']
        }

    # Step 1: Validate core totals
    total_assets = get_value('total_assets')
    total_liabilities_and_total_equity = get_value('total_liabilities_and_total_equity')

    if total_assets is None or total_liabilities_and_total_equity is None:
        return None, 'failed: missing total_assets or total_liabilities_and_total_equity'

    if not is_equal_within_tolerance(total_assets, total_liabilities_and_total_equity):
        return None, f'failed: total_assets ({total_assets}) != total_liabilities_and_total_equity ({total_liabilities_and_total_equity})'

    # Step 2: Determine total_liabilities
    total_liabilities = get_value('total_liabilities')
    total_equity = get_value('total_equity')
    total_stockholders_equity = get_value('total_stockholders_equity')

    # Must have at least one of: total_liabilities, total_equity, total_stockholders_equity
    if total_liabilities is None and total_equity is None and total_stockholders_equity is None:
        return None, 'failed: missing all of total_liabilities, total_equity, and total_stockholders_equity'

    # Calculate total_liabilities if missing
    if total_liabilities is None:
        # Prefer total_equity over total_stockholders_equity (includes minority interests)
        equity_to_use = total_equity if total_equity is not None else total_stockholders_equity
        total_liabilities_periods = {}
        tl_and_e_periods = get_period_values('total_liabilities_and_total_equity')
        equity_periods = get_period_values('total_equity' if total_equity is not None else 'total_stockholders_equity')

        for period in tl_and_e_periods:
            if period in equity_periods:
                total_liabilities_periods[period] = tl_and_e_periods[period] - equity_periods[period]

        set_value('total_liabilities', total_liabilities_periods)
        total_liabilities = get_value('total_liabilities')

    # Step 3: Calculate current/non-current splits
    total_current_assets = get_value('total_current_assets')
    total_current_liabilities = get_value('total_current_liabilities')

    # Note: Previously we skipped if these were missing, but we need to generate
    # balance sheets regardless to investigate duplicate tag issues
    # if total_current_assets is None or total_current_liabilities is None:
    #     return None, 'failed: missing total_current_assets or total_current_liabilities'

    # Calculate total_non_current_assets if missing
    if 'total_non_current_assets' not in standardized:
        tl_nca_periods = {}
        ta_periods = get_period_values('total_assets')
        tca_periods = get_period_values('total_current_assets')

        for period in ta_periods:
            if period in tca_periods:
                tl_nca_periods[period] = ta_periods[period] - tca_periods[period]

        set_value('total_non_current_assets', tl_nca_periods)

    # Calculate total_non_current_liabilities if missing
    if 'total_non_current_liabilities' not in standardized:
        tl_ncl_periods = {}
        tl_periods = get_period_values('total_liabilities')
        tcl_periods = get_period_values('total_current_liabilities')

        for period in tl_periods:
            if period in tcl_periods:
                tl_ncl_periods[period] = tl_periods[period] - tcl_periods[period]

        set_value('total_non_current_liabilities', tl_ncl_periods)

    # Step 4: Calculate residual other_* items
    # Define control items to exclude from summation
    control_fields = {
        'total_assets', 'total_liabilities_and_total_equity', 'total_liabilities',
        'total_equity', 'total_stockholders_equity', 'total_current_assets',
        'total_non_current_assets', 'total_current_liabilities', 'total_non_current_liabilities'
    }

    # Define sections and their corresponding total/other fields
    sections = {
        'other_current_assets': ('total_current_assets', [
            'cash_and_cash_equivalents', 'cash_and_short_term_investments', 'cash_cash_equivalent_and_restricted_cash',
            'short_term_investments', 'account_receivables_net', 'other_receivables', 'inventory', 'prepaids'
        ]),
        'other_non_current_assets': ('total_non_current_assets', [
            'property_plant_equipment_net', 'finance_lease_right_of_use_assets',
            'operating_lease_right_of_use_assets', 'lease_assets', 'long_term_investments', 'goodwill',
            'intangible_assets', 'goodwill_and_intangible_assets', 'deferred_tax_assets'
        ]),
        'other_current_liabilities': ('total_current_liabilities', [
            'account_payables', 'accrued_payroll', 'accrued_expenses', 'short_term_debt',
            'deferred_revenue', 'tax_payables', 'dividends_payable', 'finance_lease_obligations_current',
            'operating_lease_obligations_current', 'lease_obligation_current'
        ]),
        'other_non_current_liabilities': ('total_non_current_liabilities', [
            'long_term_debt', 'pension_and_postretirement_benefits', 'deferred_revenue_non_current',
            'deferred_tax_liabilities_non_current', 'tax_payables_non_current', 'finance_lease_obligations_non_current',
            'operating_lease_obligations_non_current', 'lease_obligation_non_current', 'commitments_and_contingencies'
        ])
    }

    # Equity section: conditional based on whether total_equity exists
    equity_items = ['common_stock', 'preferred_stock', 'additional_paid_in_capital', 'treasury_stock',
                    'retained_earnings', 'accumulated_other_comprehensive_income_loss']

    # If total_equity exists, NCI items are between total_stockholders_equity and total_equity (don't include in list)
    # If total_equity doesn't exist, NCI items are part of total_stockholders_equity (include in list)
    if 'total_equity' not in standardized:
        equity_items.extend(['minority_interest', 'redeemable_non_controlling_interests'])

    sections['other_total_stockholders_equity'] = ('total_stockholders_equity', equity_items)

    # Calculate residuals for each section
    for other_field, (total_field, item_fields) in sections.items():
        total_periods = get_period_values(total_field)
        if not total_periods:
            continue

        residual_periods = {}
        for period in total_periods:
            total_val = total_periods[period]
            sum_val = 0

            # Sum all mapped items in this section
            for item_field in item_fields:
                if item_field in standardized:
                    item_periods = get_period_values(item_field)
                    if period in item_periods:
                        # Treasury stock is contra-equity: subtract it from sum (so it adds to residual)
                        if item_field == 'treasury_stock' and other_field == 'other_total_stockholders_equity':
                            sum_val -= item_periods[period]
                        else:
                            sum_val += item_periods[period]

            # Calculate residual
            residual = total_val - sum_val
            if abs(residual) > 0.01:  # Only include if non-zero (accounting for rounding)
                residual_periods[period] = residual

        # Set residual if non-empty
        if residual_periods:
            set_value(other_field, residual_periods)

    return standardized, 'success'


def get_balance_sheet_structure():
    """Define standardized balance sheet structure with sections"""
    return [
        {'type': 'major_section', 'label': 'ASSETS'},
        {'type': 'section_header', 'label': 'Current Assets'},
        {'type': 'item', 'field': 'cash_and_cash_equivalents', 'label': 'Cash and cash equivalents', 'indent': 1},
        {'type': 'item', 'field': 'cash_and_short_term_investments', 'label': 'Cash and short-term investments', 'indent': 1},
        {'type': 'item', 'field': 'cash_cash_equivalent_and_restricted_cash', 'label': 'Cash, cash equivalents and restricted cash', 'indent': 1},
        {'type': 'item', 'field': 'short_term_investments', 'label': 'Short-term investments', 'indent': 1},
        {'type': 'item', 'field': 'account_receivables_net', 'label': 'Accounts receivable, net', 'indent': 1},
        {'type': 'item', 'field': 'other_receivables', 'label': 'Other receivables', 'indent': 1},
        {'type': 'item', 'field': 'inventory', 'label': 'Inventory', 'indent': 1},
        {'type': 'item', 'field': 'prepaids', 'label': 'Prepaid expenses', 'indent': 1},
        {'type': 'item', 'field': 'other_current_assets', 'label': 'Other current assets', 'indent': 1},
        {'type': 'subtotal', 'field': 'total_current_assets', 'label': 'Total Current Assets'},
        {'type': 'section_header', 'label': 'Non-Current Assets'},
        {'type': 'item', 'field': 'property_plant_equipment_net', 'label': 'Property, plant and equipment, net', 'indent': 1},
        {'type': 'item', 'field': 'finance_lease_right_of_use_assets', 'label': 'Finance lease right-of-use assets', 'indent': 1},
        {'type': 'item', 'field': 'operating_lease_right_of_use_assets', 'label': 'Operating lease right-of-use assets', 'indent': 1},
        {'type': 'item', 'field': 'lease_assets', 'label': 'Lease right-of-use assets', 'indent': 1},
        {'type': 'item', 'field': 'long_term_investments', 'label': 'Long-term investments', 'indent': 1},
        {'type': 'item', 'field': 'goodwill', 'label': 'Goodwill', 'indent': 1},
        {'type': 'item', 'field': 'intangible_assets', 'label': 'Intangible assets, net', 'indent': 1},
        {'type': 'item', 'field': 'goodwill_and_intangible_assets', 'label': 'Goodwill and intangible assets', 'indent': 1},
        {'type': 'item', 'field': 'deferred_tax_assets', 'label': 'Deferred tax assets', 'indent': 1},
        {'type': 'item', 'field': 'other_non_current_assets', 'label': 'Other non-current assets', 'indent': 1},
        {'type': 'subtotal', 'field': 'total_non_current_assets', 'label': 'Total Non-Current Assets'},
        {'type': 'subtotal', 'field': 'total_assets', 'label': 'TOTAL ASSETS'},
        {'type': 'blank'},
        {'type': 'major_section', 'label': 'LIABILITIES AND STOCKHOLDERS\' EQUITY'},
        {'type': 'section_header', 'label': 'Current Liabilities'},
        {'type': 'item', 'field': 'short_term_debt', 'label': 'Short-term debt', 'indent': 1},
        {'type': 'item', 'field': 'account_payables', 'label': 'Accounts payable', 'indent': 1},
        {'type': 'item', 'field': 'accrued_payroll', 'label': 'Accrued compensation', 'indent': 1},
        {'type': 'item', 'field': 'accrued_expenses', 'label': 'Accrued expenses', 'indent': 1},
        {'type': 'item', 'field': 'deferred_revenue', 'label': 'Deferred revenue', 'indent': 1},
        {'type': 'item', 'field': 'tax_payables', 'label': 'Income taxes payable', 'indent': 1},
        {'type': 'item', 'field': 'dividends_payable', 'label': 'Dividends payable', 'indent': 1},
        {'type': 'item', 'field': 'finance_lease_obligations_current', 'label': 'Finance lease liabilities - current', 'indent': 1},
        {'type': 'item', 'field': 'operating_lease_obligations_current', 'label': 'Operating lease liabilities - current', 'indent': 1},
        {'type': 'item', 'field': 'lease_obligation_current', 'label': 'Lease liabilities - current', 'indent': 1},
        {'type': 'item', 'field': 'other_payables', 'label': 'Other payables', 'indent': 1},
        {'type': 'item', 'field': 'other_current_liabilities', 'label': 'Other current liabilities', 'indent': 1},
        {'type': 'subtotal', 'field': 'total_current_liabilities', 'label': 'Total Current Liabilities'},
        {'type': 'section_header', 'label': 'Non-Current Liabilities'},
        {'type': 'item', 'field': 'long_term_debt', 'label': 'Long-term debt', 'indent': 1},
        {'type': 'item', 'field': 'pension_and_postretirement_benefits', 'label': 'Pension and postretirement benefits', 'indent': 1},
        {'type': 'item', 'field': 'deferred_revenue_non_current', 'label': 'Deferred revenue, non-current', 'indent': 1},
        {'type': 'item', 'field': 'deferred_tax_liabilities_non_current', 'label': 'Deferred tax liabilities', 'indent': 1},
        {'type': 'item', 'field': 'tax_payables_non_current', 'label': 'Income tax payable, non-current', 'indent': 1},
        {'type': 'item', 'field': 'finance_lease_obligations_non_current', 'label': 'Finance lease liabilities - non-current', 'indent': 1},
        {'type': 'item', 'field': 'operating_lease_obligations_non_current', 'label': 'Operating lease liabilities - non-current', 'indent': 1},
        {'type': 'item', 'field': 'lease_obligation_non_current', 'label': 'Lease liabilities - non-current', 'indent': 1},
        {'type': 'item', 'field': 'commitments_and_contingencies', 'label': 'Commitments and contingencies', 'indent': 1},
        {'type': 'item', 'field': 'other_non_current_liabilities', 'label': 'Other non-current liabilities', 'indent': 1},
        {'type': 'subtotal', 'field': 'total_non_current_liabilities', 'label': 'Total Non-Current Liabilities'},
        {'type': 'subtotal', 'field': 'total_liabilities', 'label': 'Total Liabilities'},
        {'type': 'section_header', 'label': 'Stockholders\' Equity'},
        {'type': 'item', 'field': 'common_stock', 'label': 'Common stock', 'indent': 1},
        {'type': 'item', 'field': 'preferred_stock', 'label': 'Preferred stock', 'indent': 1},
        {'type': 'item', 'field': 'additional_paid_in_capital', 'label': 'Additional paid-in capital', 'indent': 1},
        {'type': 'item', 'field': 'treasury_stock', 'label': 'Treasury stock', 'indent': 1},
        {'type': 'item', 'field': 'retained_earnings', 'label': 'Retained earnings', 'indent': 1},
        {'type': 'item', 'field': 'accumulated_other_comprehensive_income_loss', 'label': 'Accumulated other comprehensive income (loss)', 'indent': 1},
        {'type': 'item', 'field': 'other_total_stockholders_equity', 'label': 'Other stockholders\' equity', 'indent': 1},
        {'type': 'subtotal', 'field': 'total_stockholders_equity', 'label': 'Total Stockholders\' Equity'},
        {'type': 'item', 'field': 'minority_interest', 'label': 'Noncontrolling interests', 'indent': 1},
        {'type': 'item', 'field': 'redeemable_non_controlling_interests', 'label': 'Redeemable noncontrolling interests', 'indent': 1},
        {'type': 'subtotal', 'field': 'total_equity', 'label': 'Total Equity'},
        {'type': 'total', 'field': 'total_liabilities_and_total_equity', 'label': 'TOTAL LIABILITIES AND STOCKHOLDERS\' EQUITY'},
    ]


def get_income_statement_structure():
    """Define standardized income statement structure (CSV lines 54-81)"""
    return [
        {'type': 'item', 'field': 'revenue', 'label': 'Revenue'},
        {'type': 'item', 'field': 'cost_of_revenue', 'label': 'Cost of revenue', 'indent': 1},
        {'type': 'subtotal', 'field': 'gross_profit', 'label': 'Gross Profit'},
        {'type': 'blank'},
        {'type': 'section_header', 'label': 'Operating Expenses'},
        {'type': 'item', 'field': 'research_and_development_expenses', 'label': 'Research and development', 'indent': 1},
        {'type': 'item', 'field': 'sales_and_marketing_expenses', 'label': 'Sales and marketing', 'indent': 1},
        {'type': 'item', 'field': 'general_and_administrative_expenses', 'label': 'General and administrative', 'indent': 1},
        {'type': 'item', 'field': 'selling_general_and_administrative_expenses', 'label': 'Selling, general and administrative', 'indent': 1},
        {'type': 'item', 'field': 'depreciation_and_amortization', 'label': 'Depreciation and amortization', 'indent': 1},
        {'type': 'item', 'field': 'other_expenses', 'label': 'Other expenses', 'indent': 1},
        {'type': 'item', 'field': 'operating_expenses', 'label': 'Total operating expenses', 'indent': 1},
        {'type': 'item', 'field': 'cost_and_expenses', 'label': 'Total costs and expenses', 'indent': 1},
        {'type': 'subtotal', 'field': 'operating_income', 'label': 'Operating Income'},
        {'type': 'blank'},
        {'type': 'section_header', 'label': 'Non-Operating Items'},
        {'type': 'item', 'field': 'interest_income', 'label': 'Interest income', 'indent': 1},
        {'type': 'item', 'field': 'interest_expense', 'label': 'Interest expense', 'indent': 1},
        {'type': 'item', 'field': 'net_interest_income', 'label': 'Net interest income', 'indent': 1},
        {'type': 'item', 'field': 'total_other_income_expenses_net', 'label': 'Other income (expense), net', 'indent': 1},
        {'type': 'item', 'field': 'non_operating_income', 'label': 'Non-operating income', 'indent': 1},
        {'type': 'subtotal', 'field': 'income_before_tax', 'label': 'Income Before Income Taxes'},
        {'type': 'item', 'field': 'income_tax_expense', 'label': 'Income tax expense', 'indent': 1},
        {'type': 'blank'},
        {'type': 'item', 'field': 'net_income_from_continuing_operations', 'label': 'Net income from continuing operations'},
        {'type': 'item', 'field': 'net_income_from_discontinued_operations', 'label': 'Net income from discontinued operations', 'indent': 1},
        {'type': 'total', 'field': 'net_income', 'label': 'Net Income'},
        {'type': 'item', 'field': 'net_income_attributed_to_non_controlling_interests', 'label': 'Less: Net income attributed to non-controlling interests', 'indent': 1},
        {'type': 'item', 'field': 'net_income_attributable_to_controlling_interests', 'label': 'Net income attributable to controlling interests'},
        {'type': 'blank'},
        {'type': 'item', 'field': 'eps', 'label': 'Earnings per share - basic'},
        {'type': 'item', 'field': 'eps_diluted', 'label': 'Earnings per share - diluted'},
        {'type': 'item', 'field': 'weighted_average_shares_outstanding', 'label': 'Weighted average shares - basic'},
        {'type': 'item', 'field': 'weighted_average_shares_outstanding_diluted', 'label': 'Weighted average shares - diluted'},
    ]


def get_cash_flow_structure():
    """Define standardized cash flow structure (CSV lines 93-161)"""
    return [
        {'type': 'major_section', 'label': 'OPERATING ACTIVITIES'},
        {'type': 'item', 'field': 'net_income_starting_line', 'label': 'Net income', 'indent': 1},
        {'type': 'section_header', 'label': 'Adjustments to reconcile net income'},
        {'type': 'item', 'field': 'depreciation_and_amortization', 'label': 'Depreciation and amortization', 'indent': 2},
        {'type': 'item', 'field': 'depreciation', 'label': 'Depreciation', 'indent': 2},
        {'type': 'item', 'field': 'amortization', 'label': 'Amortization', 'indent': 2},
        {'type': 'item', 'field': 'stock_based_compensation', 'label': 'Stock-based compensation', 'indent': 2},
        {'type': 'item', 'field': 'deferred_income_tax', 'label': 'Deferred income tax', 'indent': 2},
        {'type': 'item', 'field': 'impairments', 'label': 'Impairments', 'indent': 2},
        {'type': 'item', 'field': 'pension_and_postretirement', 'label': 'Pension and postretirement', 'indent': 2},
        {'type': 'item', 'field': 'non_operating_expense_income', 'label': 'Non-operating expense/income', 'indent': 2},
        {'type': 'item', 'field': 'investment_gains_losses', 'label': 'Investment gains/losses', 'indent': 2},
        {'type': 'item', 'field': 'other_adjustments', 'label': 'Other adjustments', 'indent': 2},
        {'type': 'section_header', 'label': 'Changes in operating assets and liabilities'},
        {'type': 'item', 'field': 'accounts_receivables', 'label': 'Accounts receivables', 'indent': 2},
        {'type': 'item', 'field': 'vendor_receivables', 'label': 'Vendor receivables', 'indent': 2},
        {'type': 'item', 'field': 'inventory', 'label': 'Inventory', 'indent': 2},
        {'type': 'item', 'field': 'prepaids', 'label': 'Prepaids', 'indent': 2},
        {'type': 'item', 'field': 'accounts_payables', 'label': 'Accounts payables', 'indent': 2},
        {'type': 'item', 'field': 'accrued_expenses', 'label': 'Accrued expenses', 'indent': 2},
        {'type': 'item', 'field': 'unearned_revenue', 'label': 'Unearned revenue', 'indent': 2},
        {'type': 'item', 'field': 'income_taxes_payable', 'label': 'Income taxes payable', 'indent': 2},
        {'type': 'item', 'field': 'income_taxes', 'label': 'Income taxes', 'indent': 2},
        {'type': 'item', 'field': 'other_assets', 'label': 'Other assets', 'indent': 2},
        {'type': 'item', 'field': 'other_liabilities', 'label': 'Other liabilities', 'indent': 2},
        {'type': 'item', 'field': 'change_in_other_working_capital', 'label': 'Change in other working capital', 'indent': 2},
        {'type': 'item', 'field': 'other_operating_activities', 'label': 'Other operating activities', 'indent': 2},
        {'type': 'subtotal', 'field': 'net_cash_provided_by_operating_activities', 'label': 'Net Cash from Operating Activities'},
        {'type': 'blank'},
        {'type': 'major_section', 'label': 'INVESTING ACTIVITIES'},
        {'type': 'item', 'field': 'investments_in_property_plant_and_equipment', 'label': 'Investments in property, plant and equipment', 'indent': 1},
        {'type': 'item', 'field': 'proceeds_from_sales_of_ppe', 'label': 'Proceeds from sales of PP&E', 'indent': 1},
        {'type': 'item', 'field': 'acquisitions_of_business_net', 'label': 'Acquisitions of business, net', 'indent': 1},
        {'type': 'item', 'field': 'proceeds_from_divestiture', 'label': 'Proceeds from divestiture', 'indent': 1},
        {'type': 'item', 'field': 'purchases_of_investments', 'label': 'Purchases of investments', 'indent': 1},
        {'type': 'item', 'field': 'sales_maturities_of_investments', 'label': 'Sales/maturities of investments', 'indent': 1},
        {'type': 'item', 'field': 'other_aquisitons_and_investments', 'label': 'Other acquisitions and investments', 'indent': 1},
        {'type': 'item', 'field': 'other_investing_activities', 'label': 'Other investing activities', 'indent': 1},
        {'type': 'subtotal', 'field': 'net_cash_provided_by_investing_activities', 'label': 'Net Cash from Investing Activities'},
        {'type': 'blank'},
        {'type': 'major_section', 'label': 'FINANCING ACTIVITIES'},
        {'type': 'section_header', 'label': 'Debt'},
        {'type': 'item', 'field': 'short_term_debt_issuance', 'label': 'Short-term debt issuance', 'indent': 2},
        {'type': 'item', 'field': 'short_term_debt_repayment', 'label': 'Short-term debt repayment', 'indent': 2},
        {'type': 'item', 'field': 'change_in_short_term_debt_net', 'label': 'Change in short-term debt, net', 'indent': 2},
        {'type': 'item', 'field': 'long_term_net_debt_issuance', 'label': 'Long-term debt issuance', 'indent': 2},
        {'type': 'item', 'field': 'long_term_net_debt_repayment', 'label': 'Long-term debt repayment', 'indent': 2},
        {'type': 'item', 'field': 'change_in_long_term_debt_net', 'label': 'Change in long-term debt, net', 'indent': 2},
        {'type': 'item', 'field': 'term_debt_issuance', 'label': 'Debt issuance', 'indent': 2},
        {'type': 'item', 'field': 'term_debt_repayment', 'label': 'Debt repayment', 'indent': 2},
        {'type': 'item', 'field': 'change_in_term_debt', 'label': 'Change in debt, net', 'indent': 2},
        {'type': 'item', 'field': 'commercial_paper_net', 'label': 'Commercial paper, net', 'indent': 2},
        {'type': 'item', 'field': 'finance_lease_repayment', 'label': 'Finance lease repayment', 'indent': 2},
        {'type': 'item', 'field': 'financing_obligations_repayment', 'label': 'Financing obligations repayment', 'indent': 2},
        {'type': 'section_header', 'label': 'Equity'},
        {'type': 'item', 'field': 'net_stock_issuance', 'label': 'Net stock issuance', 'indent': 2},
        {'type': 'item', 'field': 'net_common_stock_issuance', 'label': 'Net common stock issuance', 'indent': 2},
        {'type': 'item', 'field': 'common_stock_issuance', 'label': 'Common stock issuance', 'indent': 2},
        {'type': 'item', 'field': 'common_stock_repurchased', 'label': 'Common stock repurchased', 'indent': 2},
        {'type': 'item', 'field': 'taxes_on_share_settlement', 'label': 'Taxes on share settlement', 'indent': 2},
        {'type': 'item', 'field': 'net_preferred_stock_issuance', 'label': 'Net preferred stock issuance', 'indent': 2},
        {'type': 'item', 'field': 'preferred_stock_issuance', 'label': 'Preferred stock issuance', 'indent': 2},
        {'type': 'item', 'field': 'proceeds_from_issuance_of_stock_sp', 'label': 'Proceeds from issuance of stock (SP)', 'indent': 2},
        {'type': 'section_header', 'label': 'Dividends'},
        {'type': 'item', 'field': 'dividends_paid', 'label': 'Dividends paid', 'indent': 2},
        {'type': 'item', 'field': 'common_dividends_paid', 'label': 'Common dividends paid', 'indent': 2},
        {'type': 'item', 'field': 'preferred_dividends_paid', 'label': 'Preferred dividends paid', 'indent': 2},
        {'type': 'item', 'field': 'issuance_costs', 'label': 'Issuance costs', 'indent': 2},
        {'type': 'item', 'field': 'other_financing_activities', 'label': 'Other financing activities', 'indent': 1},
        {'type': 'subtotal', 'field': 'net_cash_provided_by_financing_activities', 'label': 'Net Cash from Financing Activities'},
        {'type': 'blank'},
        {'type': 'item', 'field': 'effect_of_foreign_exchanges_rate_changes_on_cash', 'label': 'Effect of foreign exchange rate changes on cash'},
        {'type': 'item', 'field': 'net_change_in_cash', 'label': 'Net Change in Cash'},
        {'type': 'item', 'field': 'cash_at_beginning_of_period', 'label': 'Cash at beginning of period'},
        {'type': 'total', 'field': 'cash_at_end_of_period', 'label': 'Cash at End of Period'},
        {'type': 'blank'},
        {'type': 'section_header', 'label': 'Supplemental Disclosures'},
        {'type': 'item', 'field': 'income_taxes_paid', 'label': 'Income taxes paid', 'indent': 1},
        {'type': 'item', 'field': 'interest_paid', 'label': 'Interest paid', 'indent': 1},
    ]


def create_excel_workbook(results, company_name, ticker):
    """Create Excel workbook with all statements (reconstructed + standardized side-by-side)"""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    subtotal_font = Font(bold=True)
    total_font = Font(bold=True, size=11)

    # =========================================================================
    # METADATA SHEET - Detailed line item data
    # =========================================================================
    ws_meta = wb.create_sheet("Metadata", 0)

    # Collect all line items from all statements
    all_items = []
    for stmt_key in ['balance_sheet', 'income_statement', 'cash_flow', 'eq', 'ci']:
        if stmt_key in results:
            stmt_data = results[stmt_key]
            line_items = stmt_data.get('line_items', [])
            metadata = stmt_data.get('metadata', {})
            edgar_url = metadata.get('edgar_url', 'N/A')

            for item in line_items:
                # Add statement type and edgar_url to each item
                item_copy = item.copy()
                item_copy['stmt'] = stmt_key.upper().replace('_', ' ')[:2]
                all_items.append(item_copy)

    if all_items:
        # Define columns
        columns = ['stmt', 'line', 'inpth', 'plabel', 'tag', 'ddate', 'qtrs', 'uom',
                   'negating', 'custom', 'tlabel', 'datatype', 'iord', 'crdr']

        # Column widths
        col_widths = {'stmt': 8, 'line': 6, 'inpth': 6, 'plabel': 40, 'tag': 50,
                      'ddate': 10, 'qtrs': 6, 'uom': 8, 'negating': 10, 'custom': 8,
                      'tlabel': 50, 'datatype': 12, 'iord': 6, 'crdr': 6}

        # Headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws_meta.cell(row=1, column=col_idx, value=col_name.upper())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            col_letter = get_column_letter(col_idx)
            ws_meta.column_dimensions[col_letter].width = col_widths.get(col_name, 12)

        # Data rows
        for row_idx, item in enumerate(all_items, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = item.get(col_name, '')
                ws_meta.cell(row=row_idx, column=col_idx, value=value)

    # Define structure for each statement
    # Check if Strategy 2 was used for balance sheet
    bs_data = results.get('balance_sheet')
    use_strategy2_for_bs = bs_data and bs_data.get('strategy') == 'strategy2'

    structures = {
        'Balance Sheet': get_balance_sheet_structure_strategy2() if use_strategy2_for_bs else get_balance_sheet_structure(),
        'Income Statement': get_income_statement_structure(),
        'Cash Flow': get_cash_flow_structure()
    }

    # Conditionally reposition NCI items in Balance Sheet structure
    # If total_equity doesn't exist, move NCI items before total_stockholders_equity
    if bs_data:
        standardized = bs_data.get('standardized', {})
        if 'total_equity' not in standardized:
            # Reposition: move minority_interest and redeemable_non_controlling_interests
            # before total_stockholders_equity
            bs_structure = structures['Balance Sheet']
            nci_items = []
            other_items = []

            for item in bs_structure:
                if item.get('field') in ['minority_interest', 'redeemable_non_controlling_interests']:
                    nci_items.append(item)
                elif item.get('field') == 'total_equity':
                    # Skip total_equity if it doesn't exist
                    continue
                else:
                    other_items.append(item)

            # Rebuild structure: insert NCI items before total_stockholders_equity
            new_structure = []
            for item in other_items:
                if item.get('field') == 'total_stockholders_equity':
                    # Insert NCI items before the subtotal
                    new_structure.extend(nci_items)
                    new_structure.append(item)
                else:
                    new_structure.append(item)

            structures['Balance Sheet'] = new_structure

    # Define statement names and data
    statements_to_export = [
        ('Balance Sheet', results.get('balance_sheet')),
        ('Income Statement', results.get('income_statement')),
        ('Cash Flow', results.get('cash_flow')),
        ('Equity Statement', results.get('eq')),
        ('Comprehensive Income', results.get('ci'))
    ]

    # Create sheets for each statement
    for stmt_name, data in statements_to_export:
        if not data:
            continue

        ws = wb.create_sheet(stmt_name)

        periods = data.get('periods', [])
        num_periods = len(periods)

        # Title
        ws['A1'] = f"{company_name} ({ticker})"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = stmt_name
        ws['A2'].font = Font(bold=True, size=12)

        # Add metadata from reconstructor
        metadata = data.get('metadata', {})
        ws['A3'] = f"Filing: {metadata.get('adsh', 'N/A')}"
        ws['A4'] = "EDGAR URL:"
        ws['B4'] = metadata.get('edgar_url', 'N/A')
        ws['B4'].style = 'Hyperlink'

        row = 6

        # === LEFT SIDE: RECONSTRUCTED ===
        ws[f'A{row}'] = "AS FILED (Reconstructed)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        # Merge across all period columns
        end_col = get_column_letter(1 + num_periods)
        ws.merge_cells(f'A{row}:{end_col}{row}')
        row += 1

        # Headers for reconstructed - Line Item + all periods
        ws[f'A{row}'] = "Line Item"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill

        for col_idx, period in enumerate(periods, start=2):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}{row}'] = period.get('label', '')
            ws[f'{col_letter}{row}'].font = header_font
            ws[f'{col_letter}{row}'].fill = header_fill

        header_row = row
        row += 1

        # Reconstructed line items - show ALL periods
        recon_start_row = row
        for item in data['line_items']:
            ws[f'A{row}'] = item['plabel']
            values = item.get('values', {})
            for col_idx, period in enumerate(periods, start=2):
                col_letter = get_column_letter(col_idx)
                period_label = period.get('label', '')
                value = values.get(period_label)
                if value and not pd.isna(value):
                    ws[f'{col_letter}{row}'] = float(value)
                    ws[f'{col_letter}{row}'].number_format = '#,##0'
            row += 1

        recon_end_row = row - 1

        # === RIGHT SIDE: STANDARDIZED ===
        # Only show standardized section if we have a structure for this statement
        if stmt_name in structures and data.get('standardized'):
            # Calculate starting column for standardized (after reconstructed + separator)
            std_start_col = 2 + num_periods + 1  # Line item col + period cols + separator

            row = header_row - 1
            col_letter = get_column_letter(std_start_col)
            ws[f'{col_letter}{row}'] = "STANDARDIZED"
            ws[f'{col_letter}{row}'].font = Font(bold=True, size=12)
            # Merge across all period columns
            end_col_letter = get_column_letter(std_start_col + num_periods)
            ws.merge_cells(f'{col_letter}{row}:{end_col_letter}{row}')
            row += 1

            # Headers for standardized - Item + all periods
            col_letter = get_column_letter(std_start_col)
            ws[f'{col_letter}{row}'] = "Item"
            ws[f'{col_letter}{row}'].font = header_font
            ws[f'{col_letter}{row}'].fill = header_fill

            for col_idx, period in enumerate(periods, start=std_start_col + 1):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row}'] = period.get('label', '')
                ws[f'{col_letter}{row}'].font = header_font
                ws[f'{col_letter}{row}'].fill = header_fill

            row += 1

            # Standardized statement with structure - show ALL periods
            standardized = data.get('standardized', {})
            structure = structures.get(stmt_name, [])

            for line in structure:
                line_type = line['type']

                if line_type == 'blank':
                    row += 1
                elif line_type == 'major_section':
                    col_letter = get_column_letter(std_start_col)
                    ws[f'{col_letter}{row}'] = line['label']
                    ws[f'{col_letter}{row}'].font = Font(bold=True, size=12)
                    row += 1
                elif line_type == 'section_header':
                    col_letter = get_column_letter(std_start_col)
                    ws[f'{col_letter}{row}'] = line['label']
                    ws[f'{col_letter}{row}'].font = section_font
                    ws[f'{col_letter}{row}'].fill = section_fill
                    row += 1
                elif line_type == 'item':
                    field = line['field']
                    if field in standardized:
                        indent = line.get('indent', 0)
                        col_letter = get_column_letter(std_start_col)
                        ws[f'{col_letter}{row}'] = '  ' * indent + line['label']
                        # Show values for ALL periods
                        period_values = standardized[field].get('period_values', {})
                        for col_idx, period in enumerate(periods, start=std_start_col + 1):
                            col_letter = get_column_letter(col_idx)
                            period_label = period.get('label', '')
                            value = period_values.get(period_label)
                            if value and not pd.isna(value):
                                ws[f'{col_letter}{row}'] = float(value)
                                ws[f'{col_letter}{row}'].number_format = '#,##0'
                        row += 1
                elif line_type == 'subtotal':
                    field = line['field']
                    if field in standardized:
                        col_letter = get_column_letter(std_start_col)
                        ws[f'{col_letter}{row}'] = line['label']
                        ws[f'{col_letter}{row}'].font = subtotal_font
                        # Show values for ALL periods
                        period_values = standardized[field].get('period_values', {})
                        for col_idx, period in enumerate(periods, start=std_start_col + 1):
                            col_letter = get_column_letter(col_idx)
                            period_label = period.get('label', '')
                            value = period_values.get(period_label)
                            if value and not pd.isna(value):
                                ws[f'{col_letter}{row}'] = float(value)
                                ws[f'{col_letter}{row}'].number_format = '#,##0'
                                ws[f'{col_letter}{row}'].font = subtotal_font
                        row += 1
                elif line_type == 'total':
                    field = line['field']
                    if field in standardized:
                        col_letter = get_column_letter(std_start_col)
                        ws[f'{col_letter}{row}'] = line['label']
                        ws[f'{col_letter}{row}'].font = total_font
                        # Show values for ALL periods
                        period_values = standardized[field].get('period_values', {})
                        for col_idx, period in enumerate(periods, start=std_start_col + 1):
                            col_letter = get_column_letter(col_idx)
                            period_label = period.get('label', '')
                            value = period_values.get(period_label)
                            if value and not pd.isna(value):
                                ws[f'{col_letter}{row}'] = float(value)
                                ws[f'{col_letter}{row}'].number_format = '#,##0'
                                ws[f'{col_letter}{row}'].font = total_font
                        row += 1

        # Column widths - dynamic based on number of periods
        ws.column_dimensions['A'].width = 60  # Line Item column
        # Period columns for reconstructed section
        for col_idx in range(2, 2 + num_periods):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        # Separator column
        separator_col = get_column_letter(2 + num_periods)
        ws.column_dimensions[separator_col].width = 3
        # Standardized section columns
        std_start_col = 2 + num_periods + 1
        ws.column_dimensions[get_column_letter(std_start_col)].width = 50  # Item column
        for col_idx in range(std_start_col + 1, std_start_col + 1 + num_periods):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

    return wb


def map_financial_statements(cik, adsh, year, quarter, company_name, ticker):
    """Main function to map all three financial statements"""

    print(f"\n{'='*80}")
    print(f"MAPPING FINANCIAL STATEMENTS")
    print(f"{'='*80}")
    print(f"\nCompany: {company_name} (Ticker: {ticker}, CIK: {cik})")
    print(f"Filing: {adsh}")
    print(f"Dataset: {year}Q{quarter}")

    reconstructor = StatementReconstructor(year=year, quarter=quarter)

    results = {}

    # Map Balance Sheet
    print(f"\n📋 Reconstructing Balance Sheet...")
    bs_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type='BS')
    if bs_result and bs_result.get('line_items'):
        print(f"   ✅ {len(bs_result['line_items'])} items, {len(bs_result.get('periods', []))} periods")
        control_lines = find_bs_control_items(bs_result['line_items'])

        # Check if Strategy 2 should be used (unclassified balance sheet)
        use_strategy2 = should_use_strategy2(control_lines)

        if use_strategy2:
            print(f"   📌 Using Strategy 2 (unclassified balance sheet - missing total_current_assets or total_current_liabilities)")
            mappings, target_to_plabels = map_balance_sheet_strategy2(bs_result['line_items'], control_lines)
            standardized_base = aggregate_by_target(target_to_plabels, bs_result['line_items'])

            # Calculate residuals for Strategy 2 (other_assets, other_liabilities)
            standardized = calculate_residuals_strategy2(standardized_base, control_lines)
            status = "Strategy 2 - unclassified balance sheet"
        else:
            print(f"   📌 Using Strategy 1 (classified balance sheet)")
            mappings, target_to_plabels = map_statement('BS', bs_result['line_items'], control_lines)
            standardized_base = aggregate_by_target(target_to_plabels, bs_result['line_items'])

            # Validate and calculate residual other_* items
            standardized, status = validate_and_calculate_bs_residuals(standardized_base, control_lines, sic_code=None)
            if standardized is None:
                print(f"   ⚠️  Validation failed: {status} - exporting anyway for debugging")
                standardized = standardized_base  # Use pre-validation data
            else:
                print(f"   ✅ Validation: {status}")

        # Always export balance sheet, even if validation failed
        coverage = len(mappings) / len(bs_result['line_items']) * 100 if bs_result['line_items'] else 0
        print(f"   📊 Mapped: {len(mappings)}/{len(bs_result['line_items'])} ({coverage:.1f}%)")
        print(f"   🎯 Unique targets: {len(standardized)}")

        results['balance_sheet'] = {
            'line_items': bs_result['line_items'],
            'periods': bs_result.get('periods', []),
            'mappings': mappings,
            'standardized': standardized,
            'control_items': control_lines,
            'metadata': bs_result.get('metadata', {}),
            'validation_status': status,
            'strategy': 'strategy2' if use_strategy2 else 'strategy1'
        }

    # Map Income Statement
    print(f"\n📋 Reconstructing Income Statement...")
    is_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type='IS')
    if is_result and is_result.get('line_items'):
        print(f"   ✅ {len(is_result['line_items'])} items, {len(is_result.get('periods', []))} periods")
        control_lines = find_is_control_items(is_result['line_items'])
        mappings, target_to_plabels = map_statement('IS', is_result['line_items'], control_lines)
        standardized = aggregate_by_target(target_to_plabels, is_result['line_items'])

        coverage = len(mappings) / len(is_result['line_items']) * 100 if is_result['line_items'] else 0
        print(f"   📊 Mapped: {len(mappings)}/{len(is_result['line_items'])} ({coverage:.1f}%)")
        print(f"   🎯 Unique targets: {len(standardized)}")

        results['income_statement'] = {
            'line_items': is_result['line_items'],
            'periods': is_result.get('periods', []),
            'mappings': mappings,
            'standardized': standardized,
            'control_items': control_lines,
            'metadata': is_result.get('metadata', {})
        }

    # Map Cash Flow
    print(f"\n📋 Reconstructing Cash Flow...")
    cf_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type='CF')
    if cf_result and cf_result.get('line_items'):
        print(f"   ✅ {len(cf_result['line_items'])} items, {len(cf_result.get('periods', []))} periods")
        control_lines = find_cf_control_items(cf_result['line_items'])
        mappings, target_to_plabels = map_statement('CF', cf_result['line_items'], control_lines)
        standardized = aggregate_by_target(target_to_plabels, cf_result['line_items'])

        coverage = len(mappings) / len(cf_result['line_items']) * 100 if cf_result['line_items'] else 0
        print(f"   📊 Mapped: {len(mappings)}/{len(cf_result['line_items'])} ({coverage:.1f}%)")
        print(f"   🎯 Unique targets: {len(standardized)}")

        results['cash_flow'] = {
            'line_items': cf_result['line_items'],
            'periods': cf_result.get('periods', []),
            'mappings': mappings,
            'standardized': standardized,
            'control_items': control_lines,
            'metadata': cf_result.get('metadata', {})
        }

    # Reconstruct other available statements (EQ, CI, etc.)
    print(f"\n📋 Reconstructing other available statements...")
    for stmt_type in ['EQ', 'CI']:
        try:
            stmt_result = reconstructor.reconstruct_statement_multi_period(cik=cik, adsh=adsh, stmt_type=stmt_type)
            if stmt_result and stmt_result.get('line_items'):
                stmt_name = 'Equity Statement' if stmt_type == 'EQ' else 'Comprehensive Income'
                print(f"   ✅ {stmt_name}: {len(stmt_result['line_items'])} items, {len(stmt_result.get('periods', []))} periods")

                # Store reconstructed data (no mapping for now, just include as-is)
                results[stmt_type.lower()] = {
                    'line_items': stmt_result['line_items'],
                    'periods': stmt_result.get('periods', []),
                    'metadata': stmt_result.get('metadata', {})
                }
        except Exception as e:
            print(f"   ⚠️  {stmt_type} statement not available or error: {e}")

    # Export to Excel
    print(f"\n📊 Creating Excel workbook...")
    wb = create_excel_workbook(
        results,
        company_name,
        ticker
    )

    output_dir = Path('output/financial_statements')
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"{ticker}_{cik}_financial_statements.xlsx"

    wb.save(output_file)
    print(f"   ✅ Excel saved to: {output_file}")

    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Map all financial statements to standardized schema')
    parser.add_argument('--cik', required=True, help='Company CIK')
    parser.add_argument('--adsh', required=True, help='Filing ADSH')

    args = parser.parse_args()

    # Get company info from database
    conn = psycopg2.connect(config.get_db_connection())
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT c.company_name, c.ticker, f.source_year, f.source_quarter
        FROM companies c
        JOIN filings f ON c.cik = f.cik
        WHERE c.cik = %s AND f.adsh = %s
    """, (args.cik, args.adsh))

    info = cur.fetchone()
    cur.close()
    conn.close()

    if not info:
        print(f"❌ Filing not found: CIK {args.cik}, ADSH {args.adsh}")
        sys.exit(1)

    map_financial_statements(
        cik=args.cik,
        adsh=args.adsh,
        year=info['source_year'],
        quarter=info['source_quarter'],
        company_name=info['company_name'],
        ticker=info['ticker'] or 'N/A'
    )
